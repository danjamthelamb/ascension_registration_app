from __future__ import annotations

import base64
import html
import re
import uuid
from io import BytesIO
from datetime import date, timedelta
from pathlib import Path
from textwrap import dedent

import pandas as pd
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from ui_theme import inject_theme


from db import (
    create_admin_verification,
    create_household_verification,
    create_message_draft,
    create_gateway_test_message,
    cancel_gateway_test_message,
    delete_message_draft,
    get_admin_roster,
    get_household_references_by_email,
    get_gateway_test_messages,
    get_message_history,
    get_message_recipients,
    get_registration_by_reference,
    get_roster_groups,
    queue_message,
    init_db,
    save_registration,
    update_registration,
    update_roster_group_catechists,
    update_roster_group_classroom,
    verify_admin_code,
    verify_household_code,
)

from email_service import (
    send_admin_verification_email,
    send_household_id_recovery,
    send_registration_confirmation,
    send_update_confirmation,
    send_verification_email,
)


# ---------------------------------------------------------
# Paths
# ---------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent
LOGO_PATH = PROJECT_ROOT / "assets" / "ascension_logo.png"
FAITH_FORMATION_TERM = "2026-2027"


# ---------------------------------------------------------
# App setup
# ---------------------------------------------------------

st.set_page_config(
    page_title="Ascension Faith Formation Registration",
    page_icon="⛪",
    layout="centered",
)

inject_theme()

# ---------------------------------------------------------
# Global styling
# ---------------------------------------------------------

st.markdown(
    """
    <style>

    /* --------------------------------------------------
       General
    -------------------------------------------------- */

    .block-container {
        padding-top: 2rem;
        padding-bottom: 4rem;
    }

    /* --------------------------------------------------
       Keyed cards
    -------------------------------------------------- */

    .st-key-progress_card,
    .st-key-household_empty_card,
    .st-key-household_card,
    .st-key-children_empty_card,
    [class*="st-key-child_card_"],
    .st-key-review_ready_card,
    .st-key-review_pending_card,
    .st-key-landing_new_card,
    .st-key-landing_returning_card,
    .st-key-completion_household_id_card,
    .st-key-completion_next_steps_card,
    .st-key-review_dialog_household_card,
    [class*="st-key-review_dialog_child_card_"],
    .st-key-review_dialog_submit_card,
    .st-key-messaging_history_detail,
    [class*="st-key-roster_card_"] {
        background-color: #FFFDFC !important;
        border: 1px solid #C2B5A5 !important;
        border-radius: 12px !important;
        box-shadow:
            inset 0 0 0 1px rgba(255, 255, 255, 0.65),
            0 2px 7px rgba(32, 58, 92, 0.045) !important;
        overflow: hidden;
    }

    /* --------------------------------------------------
       Landing page
    -------------------------------------------------- */

    .landing-logo {
        text-align: center;
        margin-top: 0.25rem;
        margin-bottom: 1rem;
    }

    .landing-logo img {
        width: 125px;
        max-width: 34vw;
        height: auto;
    }

    .landing-parish {
        text-align: center;
        font-size: 0.9rem;
        font-weight: 700;
        letter-spacing: 0.11em;
        text-transform: uppercase;
        opacity: 0.72;
        margin-bottom: 0.35rem;
    }

    .landing-title {
        text-align: center;
        font-size: 2.35rem;
        font-weight: 800;
        line-height: 1.12;
        letter-spacing: -0.025em;
        margin-bottom: 0.9rem;
    }

    .landing-welcome {
        max-width: 610px;
        margin: 0 auto 2.2rem auto;
        text-align: center;
        font-size: 1.03rem;
        line-height: 1.6;
        opacity: 0.78;
    }

    .landing-section-label {
        font-size: 0.78rem;
        font-weight: 700;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        opacity: 0.55;
        margin-bottom: 0.2rem;
    }

    .landing-action-title {
        font-size: 1.08rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }

    .landing-action-description {
        font-size: 0.9rem;
        line-height: 1.4;
        opacity: 0.67;
        margin-bottom: 0.85rem;
    }

    .landing-recovery {
        text-align: center;
        margin-top: 1rem;
        margin-bottom: -0.4rem;
        font-size: 0.82rem;
        opacity: 0.55;
    }

    .landing-admin {
        text-align: center;
        margin-top: 1.3rem;
        margin-bottom: 0.4rem;
        font-size: 0.72rem;
        opacity: 0.45;
        text-transform: uppercase;
        letter-spacing: 0.1em;
    }

    /* --------------------------------------------------
       Registration workspace
    -------------------------------------------------- */

    .registration-intro {
        font-size: 1rem;
        line-height: 1.55;
        opacity: 0.75;
        margin-bottom: 1.2rem;
    }

    .progress-label {
        font-size: 0.75rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.07em;
        opacity: 0.55;
        margin-bottom: 0.2rem;
    }

    .progress-title {
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.15rem;
    }

    .progress-complete {
        font-size: 0.83rem;
        font-weight: 600;
    }

    .progress-waiting {
        font-size: 0.83rem;
        opacity: 0.55;
    }

    .progress-needed {
        font-size: 0.83rem;
        font-weight: 600;
    }

    .section-eyebrow {
        font-size: 0.76rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        opacity: 0.5;
        margin-bottom: 0.2rem;
    }

    .empty-state-title {
        font-size: 1.08rem;
        font-weight: 700;
        margin-bottom: 0.25rem;
    }

    .empty-state-copy {
        font-size: 0.9rem;
        line-height: 1.5;
        opacity: 0.68;
        margin-bottom: 0.9rem;
    }

    .review-ready-title {
        font-size: 1.05rem;
        font-weight: 700;
        margin-bottom: 0.35rem;
    }

    .review-check {
        font-size: 0.9rem;
        line-height: 1.6;
    }

    .privacy-note {
        text-align: center;
        font-size: 0.77rem;
        line-height: 1.45;
        opacity: 0.42;
        margin-top: 2.5rem;
    }

    /* --------------------------------------------------
       Review dialog
    -------------------------------------------------- */

    .review-section-title {
        font-size: 1.08rem;
        font-weight: 850;
        letter-spacing: 0.09em;
        text-transform: uppercase;
        margin-top: 0.2rem;
        margin-bottom: 1rem;
    }

    .review-person-name {
        font-size: 1.03rem;
        font-weight: 750;
        margin-bottom: 1rem;
    }

    .review-field-label {
        font-size: 0.7rem;
        font-weight: 750;
        letter-spacing: 0.07em;
        text-transform: uppercase;
        opacity: 0.48;
        margin-bottom: 0.1rem;
    }

    .review-field-value {
        font-size: 0.94rem;
        font-weight: 550;
        line-height: 1.45;
        margin-bottom: 1rem;
    }

    .review-child-name {
        font-size: 1.02rem;
        font-weight: 750;
        margin-bottom: 1rem;
    }

    .review-submit-title {
        font-size: 1rem;
        font-weight: 750;
        margin-bottom: 0.25rem;
    }

    .review-submit-copy {
        font-size: 0.9rem;
        line-height: 1.5;
        opacity: 0.68;
        margin-bottom: 0.75rem;
    }

    /* --------------------------------------------------
       Completion screen
    -------------------------------------------------- */

    .completion-logo {
        text-align: center;
        margin-bottom: 0.8rem;
    }

    .completion-logo img {
        width: 92px;
        max-width: 28vw;
        height: auto;
    }

    .completion-kicker {
        text-align: center;
        font-size: 0.72rem;
        font-weight: 800;
        letter-spacing: 0.11em;
        text-transform: uppercase;
        color: #586A82;
        opacity: 1;
        margin-bottom: 0.45rem;
    }

    .completion-title {
        text-align: center;
        font-size: 2rem;
        font-weight: 800;
        line-height: 1.15;
        color: #203A5C;
        margin-bottom: 0.55rem;
    }

    .completion-copy {
        text-align: center;
        max-width: 590px;
        margin: 0 auto 1.5rem auto;
        font-size: 0.96rem;
        line-height: 1.55;
        color: #505861;
        opacity: 1;
    }

    .completion-summary {
        background: #FFFDFC;
        border: 1px solid #D8D0C5;
        border-radius: 14px;
        padding: 1.35rem 1.5rem;
        margin: 1.25rem 0 1.15rem 0;
        text-align: center;
        box-shadow: 0 2px 8px rgba(32, 58, 92, 0.035);
    }

    .completion-id-label {
        color: #586A82;
        font-size: 0.72rem;
        font-weight: 800;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        margin-bottom: 0.55rem;
    }

    .completion-id-value {
        display: inline-block;
        background: #F2EEE6;
        color: #203A5C;
        border: 1px solid #D8D0C5;
        border-radius: 8px;
        padding: 0.7rem 1.15rem;
        font-family: "Courier New", Courier, monospace;
        font-size: 1.2rem;
        line-height: 1.1;
        font-weight: 700;
        letter-spacing: 0.055em;
        margin-bottom: 0.85rem;
    }

    .completion-id-help {
        color: #505861;
        font-size: 0.87rem;
        line-height: 1.5;
        max-width: 525px;
        margin: 0 auto;
    }

    .completion-email-success {
        display: flex;
        align-items: flex-start;
        gap: 0.6rem;
        background: #E8F1E5;
        border: 1px solid #C9DDC4;
        border-radius: 10px;
        padding: 0.85rem 1rem;
        margin-top: 1rem;
        text-align: left;
        color: #354638;
        font-size: 0.9rem;
        line-height: 1.45;
    }

    .completion-email-success strong {
        color: #2F4132;
    }

    .completion-email-check {
        color: #5F775C;
        font-weight: 800;
        line-height: 1.45;
        flex: 0 0 auto;
    }

    .completion-email-warning {
        background: #F7EEDC;
        border: 1px solid #E3D1AD;
        border-radius: 10px;
        padding: 0.85rem 1rem;
        margin-top: 1rem;
        color: #5D4A27;
        font-size: 0.9rem;
        line-height: 1.45;
        text-align: left;
    }

    .completion-next {
        background: #FFFDFC;
        border: 1px solid #D8D0C5;
        border-radius: 12px;
        padding: 1.15rem 1.3rem;
        margin: 1rem 0 1.35rem 0;
        box-shadow: 0 2px 8px rgba(32, 58, 92, 0.025);
    }

    .completion-next-title {
        color: #203A5C;
        font-size: 1rem;
        font-weight: 750;
        margin-bottom: 0.7rem;
    }

    .completion-next-item {
        color: #3F4750;
        font-size: 0.9rem;
        line-height: 1.55;
        margin-bottom: 0.3rem;
    }

    .completion-next-item:last-child {
        margin-bottom: 0;
    }

    .completion-next-check {
        color: #6A8068;
        font-weight: 800;
        margin-right: 0.35rem;
    }

    .completion-footer {
        text-align: center;
        margin-top: 1.45rem;
        font-size: 0.75rem;
        line-height: 1.55;
        color: #626A72;
        opacity: 0.85;
    }

    </style>
    """,
    unsafe_allow_html=True,
)


init_db()


# ---------------------------------------------------------
# Helpers
# ---------------------------------------------------------

def image_to_data_url(path: Path) -> str:

    suffix = path.suffix.lower()

    mime_types = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }

    mime_type = mime_types.get(
        suffix,
        "image/png",
    )

    encoded = base64.b64encode(
        path.read_bytes()
    ).decode("utf-8")

    return (
        f"data:{mime_type};"
        f"base64,{encoded}"
    )


def escape_html(value) -> str:

    if value is None:
        return ""

    return html.escape(
        str(value)
    )


def calculate_age(
    date_of_birth: date,
) -> int:

    today = date.today()

    age = (
        today.year
        - date_of_birth.year
    )

    if (
        today.month,
        today.day,
    ) < (
        date_of_birth.month,
        date_of_birth.day,
    ):
        age -= 1

    return age


def is_valid_email(
    email: str,
) -> bool:

    email = email.strip()

    pattern = (
        r"^[A-Za-z0-9._%+-]+"
        r"@[A-Za-z0-9-]+"
        r"(?:\.[A-Za-z0-9-]+)+$"
    )

    return (
        re.fullmatch(
            pattern,
            email,
        )
        is not None
    )


def normalize_phone(
    phone: str,
) -> str | None:

    digits = re.sub(
        r"\D",
        "",
        phone,
    )

    if (
        len(digits) == 11
        and digits.startswith("1")
    ):
        digits = digits[1:]

    if len(digits) != 10:
        return None

    if digits[0] in "01":
        return None

    if digits[3] in "01":
        return None

    return (
        f"({digits[:3]}) "
        f"{digits[3:6]}-"
        f"{digits[6:]}"
    )


def normalize_zip(
    zip_code: str,
) -> str | None:

    zip_code = zip_code.strip()

    if re.fullmatch(
        r"\d{5}",
        zip_code,
    ):
        return zip_code

    if re.fullmatch(
        r"\d{9}",
        zip_code,
    ):
        return (
            f"{zip_code[:5]}-"
            f"{zip_code[5:]}"
        )

    if re.fullmatch(
        r"\d{5}-\d{4}",
        zip_code,
    ):
        return zip_code

    return None


def mask_email(
    email: str,
) -> str:

    if "@" not in email:
        return email

    username, domain = (
        email.split(
            "@",
            1,
        )
    )

    if len(username) <= 1:
        masked_username = "•"

    else:
        masked_username = (
            username[0]
            + "•"
            * (
                len(username)
                - 1
            )
        )

    return (
        f"{masked_username}"
        f"@{domain}"
    )


def household_contact_complete(
    household: dict | None,
) -> bool:

    if not household:
        return False

    required_fields = [
        "parent_a_first_name",
        "parent_a_last_name",
        "parent_a_email",
        "parent_a_phone",
        "address_line_1",
        "city",
        "state",
        "zip_code",
        "emergency_contact_name",
        "emergency_contact_relationship",
        "emergency_contact_phone",
    ]

    return all(
        str(
            household.get(
                field,
                "",
            )
            or ""
        ).strip()

        for field
        in required_fields
    )


def get_admin_emails() -> set[str]:

    try:

        emails = (
            st.secrets[
                "admins"
            ][
                "emails"
            ]
        )

    except (
        KeyError,
        FileNotFoundError,
    ):
        return set()

    if isinstance(
        emails,
        str,
    ):
        emails = [
            emails
        ]

    return {
        str(email)
        .strip()
        .lower()

        for email in emails

        if str(email).strip()
    }


def is_authorized_admin(
    email: str,
) -> bool:

    return (
        email
        .strip()
        .lower()
        in get_admin_emails()
    )


def clear_verification_state() -> None:

    st.session_state.verification_reference = None
    st.session_state.verification_email = None
    st.session_state.show_existing_dialog = False


def clear_recovery_state() -> None:

    st.session_state.show_recovery_dialog = False
    st.session_state.recovery_request_sent = False


def clear_admin_login_state() -> None:

    st.session_state.admin_verification_email = None
    st.session_state.show_admin_dialog = False


def clear_admin_child_detail() -> None:

    st.session_state.admin_detail_child_id = None
    st.session_state.admin_detail_table_nonce += 1


def reset_public_registration_state() -> None:

    st.session_state.household = None
    st.session_state.children = []

    st.session_state.submitted_household_id = None
    st.session_state.submitted_household_reference = None

    st.session_state.registration_mode = None

    st.session_state.existing_household_id = None
    st.session_state.existing_household_reference = None

    st.session_state.confirmation_email_sent = None
    st.session_state.confirmation_email_address = None
    st.session_state.confirmation_email_error = None

    clear_verification_state()
    clear_recovery_state()
    clear_admin_login_state()


def sacrament_status_index(
    status: str | None,
) -> int:

    options = [
        "Select one",
        "Yes",
        "No",
        "Not sure",
    ]

    if status in options:
        return options.index(
            status
        )

    return 0


def sacrament_preparation_labels(
    child: dict,
) -> list[str]:

    labels = []

    if child.get(
        "receiving_first_communion_reconciliation",
        False,
    ):

        labels.append(
            "First Reconciliation / First Communion"
        )

    if child.get(
        "receiving_confirmation",
        False,
    ):

        labels.append(
            "Confirmation"
        )

    return labels


def sacramental_follow_up_reasons(
    child: dict,
) -> list[str]:

    reasons = []

    receiving_first_communion = (
        child.get(
            "receiving_first_communion_reconciliation",
            False,
        )
    )

    receiving_confirmation = (
        child.get(
            "receiving_confirmation",
            False,
        )
    )

    baptism_status = child.get(
        "baptism_status"
    )

    first_reconciliation_status = (
        child.get(
            "first_reconciliation_status"
        )
    )

    first_communion_status = (
        child.get(
            "first_communion_status"
        )
    )

    if receiving_first_communion:

        if baptism_status != "Yes":

            reasons.append(
                "Baptism: "
                f"{baptism_status or 'Not provided'}"
            )

    if receiving_confirmation:

        baptism_reason = (
            "Baptism: "
            f"{baptism_status or 'Not provided'}"
        )

        if (
            baptism_status != "Yes"
            and baptism_reason not in reasons
        ):

            reasons.append(
                baptism_reason
            )

        if (
            first_reconciliation_status
            != "Yes"
        ):

            reasons.append(
                "First Reconciliation: "
                f"{first_reconciliation_status or 'Not provided'}"
            )

        if (
            first_communion_status
            != "Yes"
        ):

            reasons.append(
                "First Communion: "
                f"{first_communion_status or 'Not provided'}"
            )

    return reasons


def full_name(
    first_name: str | None,
    middle_name: str | None,
    last_name: str | None,
) -> str:

    return " ".join(
        part.strip()

        for part in [
            first_name or "",
            middle_name or "",
            last_name or "",
        ]

        if part
        and part.strip()
    )


def parent_name(
    first_name: str | None,
    last_name: str | None,
) -> str:

    return " ".join(
        part.strip()

        for part in [
            first_name or "",
            last_name or "",
        ]

        if part
        and part.strip()
    )


def child_grade_label(
    grade: str,
) -> str:

    if grade == "Pre-K":
        return "Pre-K"

    if grade == "K":
        return "Kindergarten"

    return f"Grade {grade}"


def roster_group_key_for_grade(
    grade: str,
) -> str | None:

    if grade in (
        "Pre-K",
        "K",
    ):
        return "kindergarten"

    if grade in (
        "1",
        "2",
        "3",
        "4",
        "5",
    ):
        return f"grade_{grade}"

    if grade in (
        "6",
        "7",
        "8",
    ):
        return "edge"

    if grade in (
        "9",
        "10",
        "11",
        "12",
    ):
        return "life_teen"

    return None


def roster_title(
    group_key: str,
    default_name: str,
) -> str:

    titles = {
        "kindergarten":
            "Kindergarten",

        "grade_1":
            "1st Grade",

        "grade_2":
            "2nd Grade",

        "grade_3":
            "3rd Grade",

        "grade_4":
            "4th Grade",

        "grade_5":
            "5th Grade",

        "edge":
            "EDGE",

        "life_teen":
            "Life Teen",
    }

    return titles.get(
        group_key,
        default_name,
    )


def program_name_for_child(
    child: dict,
) -> str:

    group_key = (
        roster_group_key_for_grade(
            child.get(
                "grade",
                "",
            )
        )
    )

    if group_key is None:
        return "Unassigned"

    return roster_title(
        group_key,
        group_key,
    )


def build_export_dataframe(
    children: list[dict],
) -> pd.DataFrame:

    rows = []

    for child_index, child in enumerate(
        children
    ):

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        parent_a = parent_name(
            child.get(
                "parent_a_first_name"
            ),
            child.get(
                "parent_a_last_name"
            ),
        )

        parent_b = parent_name(
            child.get(
                "parent_b_first_name"
            ),
            child.get(
                "parent_b_last_name"
            ),
        )

        follow_up = (
            sacramental_follow_up_reasons(
                child
            )
        )

        rows.append(
            {
                "Household ID":
                    child.get(
                        "household_reference",
                        "",
                    ),

                "Child":
                    child_name,

                "First Name":
                    child.get(
                        "first_name",
                        "",
                    ),

                "Middle Name":
                    child.get(
                        "middle_name"
                    )
                    or "",

                "Last Name":
                    child.get(
                        "last_name",
                        "",
                    ),

                "Date of Birth":
                    child[
                        "date_of_birth"
                    ].strftime(
                        "%m/%d/%Y"
                    ),

                "Age":
                    calculate_age(
                        child[
                            "date_of_birth"
                        ]
                    ),

                "Grade":
                    child.get(
                        "grade",
                        "",
                    ),

                "School":
                    child.get(
                        "school",
                        "",
                    ),

                "Parent A":
                    parent_a,

                "Parent A Email":
                    child.get(
                        "parent_a_email",
                        "",
                    ),

                "Parent A Phone":
                    child.get(
                        "parent_a_phone",
                        "",
                    ),

                "Parent B":
                    parent_b,

                "Parent B Email":
                    child.get(
                        "parent_b_email"
                    )
                    or "",

                "Parent B Phone":
                    child.get(
                        "parent_b_phone"
                    )
                    or "",

                "Address Line 1":
                    child.get(
                        "address_line_1",
                        "",
                    ),

                "Address Line 2":
                    child.get(
                        "address_line_2"
                    )
                    or "",

                "City":
                    child.get(
                        "city",
                        "",
                    ),

                "State":
                    child.get(
                        "state",
                        "",
                    ),

                "ZIP":
                    child.get(
                        "zip_code",
                        "",
                    ),

                "Emergency Contact":
                    child.get(
                        "emergency_contact_name",
                        "",
                    ),

                "Emergency Contact Relationship":
                    child.get(
                        "emergency_contact_relationship",
                        "",
                    ),

                "Emergency Contact Phone":
                    child.get(
                        "emergency_contact_phone",
                        "",
                    ),

                "First Reconciliation / "
                "First Communion Prep":
                    (
                        "Yes"
                        if child.get(
                            "receiving_first_communion_reconciliation",
                            False,
                        )
                        else "No"
                    ),

                "Confirmation Prep":
                    (
                        "Yes"
                        if child.get(
                            "receiving_confirmation",
                            False,
                        )
                        else "No"
                    ),

                "Baptism Status":
                    child.get(
                        "baptism_status"
                    )
                    or "",

                "First Reconciliation Status":
                    child.get(
                        "first_reconciliation_status"
                    )
                    or "",

                "First Communion Status":
                    child.get(
                        "first_communion_status"
                    )
                    or "",

                "Sacramental Follow-up":
                    (
                        "Yes"
                        if follow_up
                        else "No"
                    ),

                "Follow-up Reason":
                    "; ".join(
                        follow_up
                    ),
            }
        )

    return pd.DataFrame(
        rows
    )


# ---------------------------------------------------------
# Printable roster PDF
# ---------------------------------------------------------

def build_roster_pdf(
    title: str,
    classroom: str,
    catechists: str,
    children: list[dict],
) -> bytes:

    buffer = BytesIO()

    navy = colors.HexColor(
        "#203A5C"
    )

    text_color = colors.HexColor(
        "#303A45"
    )

    soft_text = colors.HexColor(
        "#5D6670"
    )

    border_color = colors.HexColor(
        "#C8BFB4"
    )

    header_fill = colors.HexColor(
        "#EAEFF4"
    )

    alternate_fill = colors.HexColor(
        "#FAF8F4"
    )

    document = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.55 * inch,
        leftMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
        title=(
            f"Ascension Catholic Church - {title} Roster"
        ),
        author="Ascension Catholic Church",
    )

    base_styles = (
        getSampleStyleSheet()
    )

    title_style = ParagraphStyle(
        "AscensionRosterTitle",
        parent=base_styles[
            "Heading1"
        ],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=19,
        textColor=navy,
        spaceAfter=14,
    )

    meta_label_style = ParagraphStyle(
        "AscensionRosterMetaLabel",
        parent=base_styles[
            "BodyText"
        ],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=14,
        textColor=text_color,
    )

    meta_value_style = ParagraphStyle(
        "AscensionRosterMetaValue",
        parent=base_styles[
            "BodyText"
        ],
        fontName="Helvetica",
        fontSize=10.5,
        leading=14,
        textColor=text_color,
    )

    table_header_style = ParagraphStyle(
        "AscensionRosterTableHeader",
        parent=base_styles[
            "BodyText"
        ],
        fontName="Helvetica-Bold",
        fontSize=9.5,
        leading=12,
        textColor=navy,
    )

    table_cell_style = ParagraphStyle(
        "AscensionRosterTableCell",
        parent=base_styles[
            "BodyText"
        ],
        fontName="Helvetica",
        fontSize=9.5,
        leading=12,
        textColor=text_color,
    )

    story = []

    story.append(
        Paragraph(
            "Ascension Catholic Church | Roster",
            title_style,
        )
    )

    classroom_display = (
        classroom.strip()
        or "Not assigned"
    )

    catechists_display = (
        catechists.strip()
        or "Not assigned"
    )

    metadata = [
        (
            "Term",
            FAITH_FORMATION_TERM,
        ),
        (
            "Grade / Program",
            title,
        ),
        (
            "Room",
            classroom_display,
        ),
        (
            "Catechists",
            catechists_display,
        ),
        (
            "Total Enrolled",
            str(
                len(
                    children
                )
            ),
        ),
    ]

    metadata_table = Table(
        [
            [
                Paragraph(
                    f"{escape_html(label)}:",
                    meta_label_style,
                ),
                Paragraph(
                    escape_html(value),
                    meta_value_style,
                ),
            ]
            for label, value
            in metadata
        ],
        colWidths=[
            1.35 * inch,
            5.55 * inch,
        ],
        hAlign="LEFT",
    )

    metadata_table.setStyle(
        TableStyle(
            [
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    0,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8,
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    story.append(
        metadata_table
    )

    story.append(
        Spacer(
            1,
            18,
        )
    )

    table_data = [
        [
            Paragraph(
                "Child",
                table_header_style,
            ),
            Paragraph(
                "Grade",
                table_header_style,
            ),
            Paragraph(
                "School",
                table_header_style,
            ),
        ]
    ]

    for child in children:

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        table_data.append(
            [
                Paragraph(
                    escape_html(
                        child_name
                    ),
                    table_cell_style,
                ),
                Paragraph(
                    escape_html(
                        child_grade_label(
                            child.get(
                                "grade",
                                "",
                            )
                        )
                    ),
                    table_cell_style,
                ),
                Paragraph(
                    escape_html(
                        child.get(
                            "school",
                            "",
                        )
                    ),
                    table_cell_style,
                ),
            ]
        )

    roster_table = Table(
        table_data,
        colWidths=[
            3.15 * inch,
            1.15 * inch,
            2.60 * inch,
        ],
        repeatRows=1,
        hAlign="LEFT",
    )

    roster_style = [
        (
            "BACKGROUND",
            (0, 0),
            (-1, 0),
            header_fill,
        ),
        (
            "BOX",
            (0, 0),
            (-1, -1),
            0.9,
            border_color,
        ),
        (
            "INNERGRID",
            (0, 0),
            (-1, -1),
            0.45,
            border_color,
        ),
        (
            "VALIGN",
            (0, 0),
            (-1, -1),
            "MIDDLE",
        ),
        (
            "LEFTPADDING",
            (0, 0),
            (-1, -1),
            8,
        ),
        (
            "RIGHTPADDING",
            (0, 0),
            (-1, -1),
            8,
        ),
        (
            "TOPPADDING",
            (0, 0),
            (-1, -1),
            8,
        ),
        (
            "BOTTOMPADDING",
            (0, 0),
            (-1, -1),
            8,
        ),
    ]

    for row_index in range(
        2,
        len(
            table_data
        ),
        2,
    ):

        roster_style.append(
            (
                "BACKGROUND",
                (0, row_index),
                (-1, row_index),
                alternate_fill,
            )
        )

    roster_table.setStyle(
        TableStyle(
            roster_style
        )
    )

    story.append(
        roster_table
    )

    def draw_footer(
        canvas,
        doc,
    ) -> None:

        canvas.saveState()

        canvas.setStrokeColor(
            border_color
        )

        canvas.setLineWidth(
            0.5
        )

        canvas.line(
            doc.leftMargin,
            0.42 * inch,
            letter[0]
            - doc.rightMargin,
            0.42 * inch,
        )

        canvas.setFont(
            "Helvetica",
            7.5,
        )

        canvas.setFillColor(
            soft_text
        )

        canvas.drawString(
            doc.leftMargin,
            0.25 * inch,
            "Ascension Catholic Church - Faith Formation",
        )

        canvas.drawRightString(
            letter[0]
            - doc.rightMargin,
            0.25 * inch,
            f"Page {doc.page}",
        )

        canvas.restoreState()

    document.build(
        story,
        onFirstPage=draw_footer,
        onLaterPages=draw_footer,
    )

    pdf_data = (
        buffer.getvalue()
    )

    buffer.close()

    return pdf_data


# ---------------------------------------------------------
# Attendance sheet
# ---------------------------------------------------------

def get_next_sundays(
    start_date: date,
    count: int = 8,
) -> list[date]:

    days_until_sunday = (
        6 - start_date.weekday()
    ) % 7

    first_sunday = (
        start_date
        + timedelta(
            days=days_until_sunday
        )
    )

    return [
        first_sunday
        + timedelta(
            weeks=index
        )
        for index in range(
            count
        )
    ]


def build_attendance_sheet(
    title: str,
    classroom: str,
    catechists: str,
    children: list[dict],
) -> bytes:

    report_date = date.today()

    sundays = get_next_sundays(
        report_date,
        count=8,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"

    navy = "203A5C"
    light_header = "EAEFF4"
    alternate_fill = "FAF8F4"
    border_color = "C8BFB4"

    thin_side = Side(
        style="thin",
        color=border_color,
    )

    cell_border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    # -----------------------------------------------------
    # Column layout
    # -----------------------------------------------------

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 11

    for column_number in range(
        3,
        11,
    ):

        column_letter = get_column_letter(
            column_number
        )

        sheet.column_dimensions[
            column_letter
        ].width = 10

    # -----------------------------------------------------
    # Title
    # -----------------------------------------------------

    sheet.merge_cells(
        "A1:J1"
    )

    title_cell = sheet["A1"]

    title_cell.value = (
        "Ascension Catholic Church  |  "
        "Attendance Sheet"
    )

    title_cell.font = Font(
        size=16,
        bold=True,
        color=navy,
    )

    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    sheet.row_dimensions[1].height = 28

    # -----------------------------------------------------
    # Roster information
    # -----------------------------------------------------

    sheet["A3"] = "Term:"

    sheet.merge_cells(
        "B3:C3"
    )

    sheet["B3"] = (
        FAITH_FORMATION_TERM
    )

    sheet.merge_cells(
        "D3:E3"
    )

    sheet["D3"] = (
        "Grade / Program:"
    )

    sheet.merge_cells(
        "F3:G3"
    )

    sheet["F3"] = title

    sheet["H3"] = "Room:"

    sheet.merge_cells(
        "I3:J3"
    )

    sheet["I3"] = (
        classroom.strip()
        or "Not assigned"
    )

    sheet["A4"] = "Catechists:"

    sheet.merge_cells(
        "B4:J4"
    )

    sheet["B4"] = (
        catechists.strip()
        or "Not assigned"
    )

    for cell_coordinate in (
        "A3",
        "D3",
        "H3",
        "A4",
    ):

        sheet[
            cell_coordinate
        ].font = Font(
            bold=True,
            color=navy,
        )

    for cell_coordinate in (
        "A3",
        "B3",
        "D3",
        "F3",
        "H3",
        "I3",
        "A4",
        "B4",
    ):

        sheet[
            cell_coordinate
        ].alignment = Alignment(
            vertical="center",
        )

    # -----------------------------------------------------
    # Catechist attendance
    # -----------------------------------------------------

    catechist_header_row = 6

    sheet.cell(
        row=catechist_header_row,
        column=1,
        value="Catechists",
    )

    for index, sunday in enumerate(
        sundays,
        start=3,
    ):

        cell = sheet.cell(
            row=catechist_header_row,
            column=index,
            value=sunday,
        )

        cell.number_format = "m/d"

    for column_number in range(
        1,
        11,
    ):

        cell = sheet.cell(
            row=catechist_header_row,
            column=column_number,
        )

        cell.font = Font(
            bold=True,
            color=navy,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=light_header,
        )

        cell.border = cell_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.cell(
        row=catechist_header_row,
        column=1,
    ).alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    catechist_names = [
        name.strip()
        for name in catechists.split(",")
        if name.strip()
    ]

    if not catechist_names:
        catechist_names = [
            ""
        ]

    current_row = (
        catechist_header_row + 1
    )

    for catechist_name in catechist_names:

        sheet.cell(
            row=current_row,
            column=1,
            value=catechist_name,
        )

        for column_number in range(
            1,
            11,
        ):

            cell = sheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.border = cell_border

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_number == 1
                    else "center"
                ),
                vertical="center",
            )

        sheet.row_dimensions[
            current_row
        ].height = 22

        current_row += 1

    # -----------------------------------------------------
    # Student attendance
    # -----------------------------------------------------

    student_header_row = (
        current_row + 2
    )

    sheet.cell(
        row=student_header_row,
        column=1,
        value="Student Name",
    )

    sheet.cell(
        row=student_header_row,
        column=2,
        value="Grade",
    )

    for index, sunday in enumerate(
        sundays,
        start=3,
    ):

        cell = sheet.cell(
            row=student_header_row,
            column=index,
            value=sunday,
        )

        cell.number_format = "m/d"

    for column_number in range(
        1,
        11,
    ):

        cell = sheet.cell(
            row=student_header_row,
            column=column_number,
        )

        cell.font = Font(
            bold=True,
            color=navy,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=light_header,
        )

        cell.border = cell_border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.cell(
        row=student_header_row,
        column=1,
    ).alignment = Alignment(
        horizontal="left",
        vertical="center",
    )

    current_row = (
        student_header_row + 1
    )

    for child in children:

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        sheet.cell(
            row=current_row,
            column=1,
            value=child_name,
        )

        sheet.cell(
            row=current_row,
            column=2,
            value=child.get(
                "grade",
                "",
            ),
        )

        for column_number in range(
            1,
            11,
        ):

            cell = sheet.cell(
                row=current_row,
                column=column_number,
            )

            cell.border = cell_border

            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_number == 1
                    else "center"
                ),
                vertical="center",
            )

        if (
            current_row
            % 2 == 0
        ):

            for column_number in range(
                1,
                11,
            ):

                sheet.cell(
                    row=current_row,
                    column=column_number,
                ).fill = PatternFill(
                    fill_type="solid",
                    fgColor=alternate_fill,
                )

        sheet.row_dimensions[
            current_row
        ].height = 23

        current_row += 1

    # -----------------------------------------------------
    # Workbook / print settings
    # -----------------------------------------------------

    # Deliberately leave panes unfrozen so the top metadata
    # does not remain locked while scrolling the student list.
    sheet.freeze_panes = None

    sheet.sheet_view.showGridLines = False

    sheet.page_setup.orientation = (
        "landscape"
    )

    sheet.page_setup.paperSize = (
        sheet.PAPERSIZE_LETTER
    )

    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    sheet.print_options.horizontalCentered = (
        True
    )

    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.4
    sheet.page_margins.bottom = 0.4

    sheet.print_area = (
        f"A1:J{max(current_row - 1, student_header_row)}"
    )

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output.getvalue()



# ---------------------------------------------------------
# Upcoming birthdays
# ---------------------------------------------------------

def get_upcoming_birthdays(
    roster: list[dict],
) -> pd.DataFrame:
    """
    Return upcoming birthdays for all registered children.

    Groups:
    - Pre-K through Grade 5 -> PSR
    - Grades 6-8 -> EDGE
    - Grades 9-12 -> Life Teen

    Includes:
    - birthdays still remaining in the current month
    - all birthdays in the next calendar month

    Results are ordered by ministry group:
    PSR, EDGE, Life Teen, then by birthday date.
    """

    today = date.today()

    current_month = today.month

    if current_month == 12:
        next_month = 1
        next_month_year = today.year + 1
    else:
        next_month = current_month + 1
        next_month_year = today.year

    group_order = {
        "PSR": 0,
        "EDGE": 1,
        "Life Teen": 2,
    }

    rows = []

    for child in roster:

        grade = str(
            child.get(
                "grade",
                "",
            )
            or ""
        )

        if grade in (
            "Pre-K",
            "K",
            "1",
            "2",
            "3",
            "4",
            "5",
        ):
            ministry_group = "PSR"

        elif grade in (
            "6",
            "7",
            "8",
        ):
            ministry_group = "EDGE"

        elif grade in (
            "9",
            "10",
            "11",
            "12",
        ):
            ministry_group = "Life Teen"

        else:
            # Ignore records without a recognized grade.
            continue

        date_of_birth = child.get(
            "date_of_birth"
        )

        if not isinstance(
            date_of_birth,
            date,
        ):
            continue

        if date_of_birth.month == current_month:
            birthday_year = today.year

        elif date_of_birth.month == next_month:
            birthday_year = next_month_year

        else:
            continue

        try:

            upcoming_birthday = date(
                birthday_year,
                date_of_birth.month,
                date_of_birth.day,
            )

        except ValueError:
            # Treat Feb. 29 birthdays as Feb. 28 in
            # non-leap years for this planning list.
            upcoming_birthday = date(
                birthday_year,
                2,
                28,
            )

        # For the current month, don't include birthdays
        # that have already passed.
        if (
            date_of_birth.month == current_month
            and upcoming_birthday < today
        ):
            continue

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        rows.append(
            {
                "_group_order":
                    group_order[
                        ministry_group
                    ],

                "_sort_date":
                    upcoming_birthday,

                "Group":
                    ministry_group,

                "Child":
                    child_name,

                "Birthday":
                    upcoming_birthday.strftime(
                        "%m/%d/%Y"
                    ),

                "Turning":
                    birthday_year
                    - date_of_birth.year,

                "Grade":
                    child_grade_label(
                        grade
                    ),

                "School":
                    child.get(
                        "school",
                        "",
                    )
                    or "",
            }
        )

    rows.sort(
        key=lambda row: (
            row["_group_order"],
            row["_sort_date"],
            row["Child"].casefold(),
        )
    )

    for row in rows:

        row.pop(
            "_group_order",
            None,
        )

        row.pop(
            "_sort_date",
            None,
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Group",
            "Child",
            "Birthday",
            "Turning",
            "Grade",
            "School",
        ],
    )



# ---------------------------------------------------------
# Household messaging helpers
# ---------------------------------------------------------

def messaging_audience_matches(
    child: dict,
    audiences: list[str],
) -> bool:

    if not audiences:
        return False

    if "All Households" in audiences:
        return True

    grade = str(
        child.get(
            "grade",
            "",
        )
        or ""
    )

    matches = []

    if "PSR" in audiences:

        matches.append(
            grade in (
                "Pre-K",
                "K",
                "1",
                "2",
                "3",
                "4",
                "5",
            )
        )

    if "EDGE" in audiences:

        matches.append(
            grade in (
                "6",
                "7",
                "8",
            )
        )

    if "Life Teen" in audiences:

        matches.append(
            grade in (
                "9",
                "10",
                "11",
                "12",
            )
        )

    if "Confirmation" in audiences:

        matches.append(
            bool(
                child.get(
                    "receiving_confirmation",
                    False,
                )
            )
        )

    if "First Communion" in audiences:

        matches.append(
            bool(
                child.get(
                    "receiving_first_communion_reconciliation",
                    False,
                )
            )
        )

    return any(
        matches
    )


def household_display_name(
    child: dict,
) -> str:

    parent_a_first = (
        child.get(
            "parent_a_first_name",
            "",
        )
        or ""
    ).strip()

    parent_a_last = (
        child.get(
            "parent_a_last_name",
            "",
        )
        or ""
    ).strip()

    parent_b_first = (
        child.get(
            "parent_b_first_name",
            "",
        )
        or ""
    ).strip()

    parent_b_last = (
        child.get(
            "parent_b_last_name",
            "",
        )
        or ""
    ).strip()

    parent_a = parent_name(
        parent_a_first,
        parent_a_last,
    )

    parent_b = parent_name(
        parent_b_first,
        parent_b_last,
    )

    if (
        parent_a
        and parent_b
        and parent_a_last
        and parent_b_last
        and parent_a_last.casefold()
        == parent_b_last.casefold()
    ):

        return (
            f"{parent_a_first} & "
            f"{parent_b_first} "
            f"{parent_a_last}"
        ).strip()

    if parent_a and parent_b:
        return (
            f"{parent_a} & "
            f"{parent_b}"
        )

    return (
        parent_a
        or parent_b
        or child.get(
            "household_reference",
            "Household",
        )
    )


def resolve_household_message_recipients(
    roster: list[dict],
    audiences: list[str],
) -> tuple[
    pd.DataFrame,
    pd.DataFrame,
    dict,
]:

    matched_children = [
        child

        for child in roster

        if messaging_audience_matches(
            child,
            audiences,
        )
    ]

    households: dict[str, dict] = {}

    for child in matched_children:

        household_reference = str(
            child.get(
                "household_reference",
                "",
            )
            or ""
        ).strip()

        if not household_reference:
            continue

        if household_reference not in households:

            households[
                household_reference
            ] = {
                "Household":
                    household_display_name(
                        child
                    ),

                "Household ID":
                    household_reference,

                "Parent A":
                    parent_name(
                        child.get(
                            "parent_a_first_name"
                        ),
                        child.get(
                            "parent_a_last_name"
                        ),
                    ),

                "Parent A Phone":
                    (
                        child.get(
                            "parent_a_phone"
                        )
                        or ""
                    ).strip(),

                "Parent B":
                    parent_name(
                        child.get(
                            "parent_b_first_name"
                        ),
                        child.get(
                            "parent_b_last_name"
                        ),
                    ),

                "Parent B Phone":
                    (
                        child.get(
                            "parent_b_phone"
                        )
                        or ""
                    ).strip(),

                "_children":
                    [],
            }

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        if (
            child_name
            and child_name
            not in households[
                household_reference
            ][
                "_children"
            ]
        ):

            households[
                household_reference
            ][
                "_children"
            ].append(
                child_name
            )

    recipient_rows = []
    missing_rows = []

    parent_a_count = 0
    parent_b_count = 0

    for household in households.values():

        parent_a_phone = (
            household[
                "Parent A Phone"
            ]
        )

        parent_b_phone = (
            household[
                "Parent B Phone"
            ]
        )

        children_text = ", ".join(
            household[
                "_children"
            ]
        )

        if parent_a_phone:

            parent_a_count += 1

            recipient_rows.append(
                {
                    "Household":
                        household[
                            "Household"
                        ],

                    "Household ID":
                        household[
                            "Household ID"
                        ],

                    "Contact":
                        household[
                            "Parent A"
                        ]
                        or "Parent / Guardian A",

                    "Phone":
                        parent_a_phone,

                    "Using":
                        "Parent A",

                    "Children":
                        children_text,
                }
            )

        elif parent_b_phone:

            parent_b_count += 1

            recipient_rows.append(
                {
                    "Household":
                        household[
                            "Household"
                        ],

                    "Household ID":
                        household[
                            "Household ID"
                        ],

                    "Contact":
                        household[
                            "Parent B"
                        ]
                        or "Parent / Guardian B",

                    "Phone":
                        parent_b_phone,

                    "Using":
                        "Parent B fallback",

                    "Children":
                        children_text,
                }
            )

        else:

            missing_rows.append(
                {
                    "Household":
                        household[
                            "Household"
                        ],

                    "Household ID":
                        household[
                            "Household ID"
                        ],

                    "Children":
                        children_text,

                    "Issue":
                        "No parent phone number",
                }
            )

    recipient_df = pd.DataFrame(
        recipient_rows,
        columns=[
            "Household",
            "Household ID",
            "Contact",
            "Phone",
            "Using",
            "Children",
        ],
    )

    missing_df = pd.DataFrame(
        missing_rows,
        columns=[
            "Household",
            "Household ID",
            "Children",
            "Issue",
        ],
    )

    if not recipient_df.empty:

        recipient_df = (
            recipient_df
            .sort_values(
                by=[
                    "Household",
                    "Household ID",
                ],
                kind="stable",
            )
            .reset_index(
                drop=True
            )
        )

    if not missing_df.empty:

        missing_df = (
            missing_df
            .sort_values(
                by=[
                    "Household",
                    "Household ID",
                ],
                kind="stable",
            )
            .reset_index(
                drop=True
            )
        )

    stats = {
        "matched_children":
            len(
                matched_children
            ),

        "unique_households":
            len(
                households
            ),

        "parent_a":
            parent_a_count,

        "parent_b":
            parent_b_count,

        "missing":
            len(
                missing_rows
            ),

        "recipients":
            len(
                recipient_rows
            ),
    }

    return (
        recipient_df,
        missing_df,
        stats,
    )


# ---------------------------------------------------------
# Session state
# ---------------------------------------------------------

if "household" not in st.session_state:
    st.session_state.household = None

if "children" not in st.session_state:
    st.session_state.children = []

if "submitted_household_id" not in st.session_state:
    st.session_state.submitted_household_id = None

if "submitted_household_reference" not in st.session_state:
    st.session_state.submitted_household_reference = None

if "registration_mode" not in st.session_state:
    st.session_state.registration_mode = None

if "existing_household_id" not in st.session_state:
    st.session_state.existing_household_id = None

if "existing_household_reference" not in st.session_state:
    st.session_state.existing_household_reference = None

if "verification_reference" not in st.session_state:
    st.session_state.verification_reference = None

if "verification_email" not in st.session_state:
    st.session_state.verification_email = None

if "show_existing_dialog" not in st.session_state:
    st.session_state.show_existing_dialog = False

if "show_recovery_dialog" not in st.session_state:
    st.session_state.show_recovery_dialog = False

if "recovery_request_sent" not in st.session_state:
    st.session_state.recovery_request_sent = False

if "show_admin_dialog" not in st.session_state:
    st.session_state.show_admin_dialog = False

if "admin_verification_email" not in st.session_state:
    st.session_state.admin_verification_email = None

if "admin_authenticated" not in st.session_state:
    st.session_state.admin_authenticated = False

if "admin_email" not in st.session_state:
    st.session_state.admin_email = None

if "admin_detail_child_id" not in st.session_state:
    st.session_state.admin_detail_child_id = None

if "admin_detail_table_nonce" not in st.session_state:
    st.session_state.admin_detail_table_nonce = 0

if "confirmation_email_sent" not in st.session_state:
    st.session_state.confirmation_email_sent = None

if "confirmation_email_address" not in st.session_state:
    st.session_state.confirmation_email_address = None

if "confirmation_email_error" not in st.session_state:
    st.session_state.confirmation_email_error = None


# ---------------------------------------------------------
# Admin login dialog
# ---------------------------------------------------------

@st.dialog("Admin Login")
def admin_login_dialog():

    if (
        st.session_state.admin_verification_email
        is None
    ):

        st.write(
            "Enter your authorized staff email address "
            "to receive a login code."
        )

        admin_email = st.text_input(
            "Email",
            placeholder="name@example.com",
        )

        if st.button(
            "Send Login Code",
            type="primary",
            use_container_width=True,
        ):

            admin_email = (
                admin_email
                .strip()
                .lower()
            )

            if not admin_email:

                st.error(
                    "Please enter your email address to continue."
                )
                return

            if not is_valid_email(
                admin_email
            ):

                st.error(
                    "Please enter a valid email address."
                )
                return

            try:

                st.session_state.admin_verification_email = (
                    admin_email
                )

                if is_authorized_admin(
                    admin_email
                ):

                    verification = (
                        create_admin_verification(
                            admin_email
                        )
                    )

                    send_admin_verification_email(
                        recipient=verification[
                            "email"
                        ],
                        verification_code=verification[
                            "code"
                        ],
                        expires_minutes=verification[
                            "expires_minutes"
                        ],
                    )

                st.session_state.show_admin_dialog = True
                st.rerun()

            except Exception:

                st.session_state.admin_verification_email = None

                st.error(
                    "We couldn't send a login code right now. "
                    "Please wait a moment and try again."
                )

        return

    admin_email = (
        st.session_state.admin_verification_email
    )

    st.success(
        "Check your email."
    )

    st.write(
        "If this address is authorized for staff access, "
        "we sent a 6-digit login code to:"
    )

    st.write(
        f"**{mask_email(admin_email)}**"
    )

    st.caption(
        "The code expires in 10 minutes."
    )

    verification_code = (
        st.text_input(
            "Login Code",
            max_chars=6,
            placeholder="123456",
        )
    )

    if st.button(
        "Verify & Sign In",
        type="primary",
        use_container_width=True,
    ):

        if not verification_code.strip():

            st.error(
                "Please enter the 6-digit login code."
            )
            return

        verified, status = (
            verify_admin_code(
                admin_email,
                verification_code,
            )
        )

        if not verified:

            if status == "expired":

                st.error(
                    "That login code has expired. "
                    "Send yourself a new code to continue."
                )

            elif status == "locked":

                st.error(
                    "Too many incorrect attempts were made with "
                    "that code. Please request a new one."
                )

            elif status == "no_active_code":

                st.error(
                    "That login code is no longer active. "
                    "Please request a new one."
                )

            else:

                st.error(
                    "That code doesn't match. "
                    "Please check the email and try again."
                )

            return

        if not is_authorized_admin(
            admin_email
        ):

            clear_admin_login_state()

            st.error(
                "Staff access could not be verified."
            )
            return

        st.session_state.admin_authenticated = True
        st.session_state.admin_email = admin_email

        clear_admin_login_state()
        st.rerun()

    st.divider()

    if st.button(
        "Send a New Code",
        use_container_width=True,
    ):

        try:

            if is_authorized_admin(
                admin_email
            ):

                verification = (
                    create_admin_verification(
                        admin_email
                    )
                )

                send_admin_verification_email(
                    recipient=verification[
                        "email"
                    ],
                    verification_code=verification[
                        "code"
                    ],
                    expires_minutes=verification[
                        "expires_minutes"
                    ],
                )

            st.success(
                "If this address is authorized for staff access, "
                "a new login code has been sent."
            )

        except Exception:

            st.error(
                "We couldn't send a new login code right now. "
                "Please wait a moment and try again."
            )

    if st.button(
        "Use a Different Email",
        use_container_width=True,
    ):

        st.session_state.admin_verification_email = None
        st.session_state.show_admin_dialog = True
        st.rerun()


# ---------------------------------------------------------
# Edit catechists dialog
# ---------------------------------------------------------

@st.dialog("Edit Catechists")
def edit_catechists_dialog(
    group: dict,
):

    group_key = group[
        "group_key"
    ]

    title = roster_title(
        group_key,
        group[
            "display_name"
        ],
    )

    st.subheader(
        title
    )

    st.caption(
        "Enter the catechist names as you would like "
        "them displayed on the roster."
    )

    catechists = st.text_input(
        "Catechists",
        value=group.get(
            "catechists",
            "",
        ),
        placeholder="Jane Smith, John Doe",
        key=(
            f"catechists_input_"
            f"{group_key}"
        ),
    )

    if st.button(
        "Save Catechists",
        type="primary",
        use_container_width=True,
        key=(
            f"save_catechists_"
            f"{group_key}"
        ),
    ):

        try:

            update_roster_group_catechists(
                group_key,
                catechists,
            )

            st.rerun()

        except Exception:

            st.error(
                "We couldn't save the catechist names right now. "
                "Please try again."
            )


# ---------------------------------------------------------
# Admin child detail dialog
# ---------------------------------------------------------

@st.dialog(
    "Child Details",
    width="medium",
    on_dismiss=clear_admin_child_detail,
)
def admin_child_detail_dialog(
    child: dict,
):

    child_name = full_name(
        child.get(
            "first_name"
        ),
        child.get(
            "middle_name"
        ),
        child.get(
            "last_name"
        ),
    )

    age = calculate_age(
        child[
            "date_of_birth"
        ]
    )

    program = (
        program_name_for_child(
            child
        )
    )

    st.subheader(
        child_name
    )

    st.write(
        f"**Grade:** "
        f"{child['grade']}"
    )

    st.write(
        f"**Program:** "
        f"{program}"
    )

    st.write(
        f"**School:** "
        f"{child['school']}"
    )

    st.write(
        f"**Date of Birth:** "
        f"{child['date_of_birth'].strftime('%m/%d/%Y')} "
        f"(Age {age})"
    )

    st.divider()

    st.subheader(
        "Household"
    )

    st.write(
        "**Household ID**"
    )

    st.code(
        child.get(
            "household_reference",
            "",
        ),
        language=None,
    )

    parent_a = parent_name(
        child.get(
            "parent_a_first_name"
        ),
        child.get(
            "parent_a_last_name"
        ),
    )

    st.write(
        "**Parent / Guardian A**"
    )

    st.write(
        parent_a
    )

    st.write(
        child.get(
            "parent_a_email",
            "",
        )
    )

    st.write(
        child.get(
            "parent_a_phone",
            "",
        )
    )

    parent_b = parent_name(
        child.get(
            "parent_b_first_name"
        ),
        child.get(
            "parent_b_last_name"
        ),
    )

    parent_b_email = (
        child.get(
            "parent_b_email"
        )
        or ""
    )

    parent_b_phone = (
        child.get(
            "parent_b_phone"
        )
        or ""
    )

    if (
        parent_b
        or parent_b_email
        or parent_b_phone
    ):

        st.write("")

        st.write(
            "**Parent / Guardian B**"
        )

        if parent_b:
            st.write(
                parent_b
            )

        if parent_b_email:
            st.write(
                parent_b_email
            )

        if parent_b_phone:
            st.write(
                parent_b_phone
            )

    st.write("")

    st.write(
        "**Home Address**"
    )

    st.write(
        child.get(
            "address_line_1",
            "",
        )
    )

    if child.get(
        "address_line_2"
    ):

        st.write(
            child[
                "address_line_2"
            ]
        )

    st.write(
        f"{child.get('city', '')}, "
        f"{child.get('state', '')} "
        f"{child.get('zip_code', '')}"
    )

    st.divider()

    st.subheader(
        "Emergency Contact"
    )

    emergency_name = (
        child.get(
            "emergency_contact_name"
        )
        or "Not provided"
    )

    emergency_relationship = (
        child.get(
            "emergency_contact_relationship"
        )
        or "Not provided"
    )

    emergency_phone = (
        child.get(
            "emergency_contact_phone"
        )
        or "Not provided"
    )

    st.write(
        f"**Name:** {emergency_name}"
    )

    st.write(
        f"**Relationship:** "
        f"{emergency_relationship}"
    )

    st.write(
        f"**Phone:** {emergency_phone}"
    )

    st.divider()

    st.subheader(
        "Sacrament Preparation"
    )

    preparation = (
        sacrament_preparation_labels(
            child
        )
    )

    if preparation:

        for item in preparation:

            st.write(
                f"✓ {item}"
            )

    else:

        st.caption(
            "No sacrament preparation selected."
        )

    st.subheader(
        "Sacramental History"
    )

    baptism = (
        child.get(
            "baptism_status"
        )
        or "—"
    )

    reconciliation = (
        child.get(
            "first_reconciliation_status"
        )
        or "—"
    )

    communion = (
        child.get(
            "first_communion_status"
        )
        or "—"
    )

    st.write(
        f"**Baptized:** "
        f"{baptism}"
    )

    st.write(
        f"**First Reconciliation:** "
        f"{reconciliation}"
    )

    st.write(
        f"**First Communion:** "
        f"{communion}"
    )

    follow_up_reasons = (
        sacramental_follow_up_reasons(
            child
        )
    )

    if follow_up_reasons:

        st.warning(
            "**Sacramental follow-up needed**\n\n"
            + "\n\n".join(
                f"• {reason}"
                for reason
                in follow_up_reasons
            )
        )

    st.divider()

    if st.button(
        "Close",
        type="primary",
        use_container_width=True,
    ):

        clear_admin_child_detail()
        st.rerun()


# ---------------------------------------------------------
# Household ID recovery
# ---------------------------------------------------------

@st.dialog("Recover Household ID")
def recover_household_id_dialog():

    if (
        st.session_state.recovery_request_sent
    ):

        st.success(
            "Check your email."
        )

        st.write(
            "If that email address is connected to an "
            "Ascension registration, we've sent the "
            "Household ID associated with it."
        )

        st.caption(
            "Please check your inbox and spam folder."
        )

        st.divider()

        if st.button(
            "Return to Existing Household",
            type="primary",
            use_container_width=True,
        ):

            clear_recovery_state()
            st.session_state.show_existing_dialog = True
            st.rerun()

        if st.button(
            "Close",
            use_container_width=True,
        ):

            clear_recovery_state()
            st.rerun()

        return

    st.write(
        "Enter the email address associated with your "
        "household registration."
    )

    st.caption(
        "You may use the email address listed for either "
        "parent or guardian."
    )

    email = st.text_input(
        "Email",
        placeholder="name@example.com",
    )

    if st.button(
        "Send Household ID",
        type="primary",
        use_container_width=True,
    ):

        email = email.strip()

        if not email:

            st.error(
                "Please enter your email address."
            )
            return

        if not is_valid_email(
            email
        ):

            st.error(
                "Please enter a valid email address."
            )
            return

        try:

            references = (
                get_household_references_by_email(
                    email
                )
            )

            if references:

                send_household_id_recovery(
                    recipient=email,
                    household_references=references,
                )

            st.session_state.recovery_request_sent = True
            st.session_state.show_recovery_dialog = True
            st.rerun()

        except Exception:

            st.error(
                "We couldn't process the recovery request right now. "
                "Please wait a moment and try again."
            )


# ---------------------------------------------------------
# Existing household dialog
# ---------------------------------------------------------

@st.dialog("Return to Existing Household")
def existing_household_dialog():

    if (
        st.session_state.verification_reference
        is None
    ):

        st.write(
            "Enter your Household ID to securely access "
            "your existing registration."
        )

        household_reference = (
            st.text_input(
                "Household ID",
                placeholder="ASC-XXXXXX",
            )
        )

        if st.button(
            "Send Verification Code",
            type="primary",
            use_container_width=True,
        ):

            if not household_reference.strip():

                st.error(
                    "Please enter your Household ID."
                )
                return

            try:

                verification = (
                    create_household_verification(
                        household_reference
                    )
                )

                if verification is None:

                    st.error(
                        "We couldn't find a registration with that "
                        "Household ID. Please check it and try again."
                    )
                    return

                send_verification_email(
                    recipient=verification[
                        "email"
                    ],
                    verification_code=verification[
                        "code"
                    ],
                    household_reference=verification[
                        "household_reference"
                    ],
                    expires_minutes=verification[
                        "expires_minutes"
                    ],
                )

                st.session_state.verification_reference = (
                    verification[
                        "household_reference"
                    ]
                )

                st.session_state.verification_email = (
                    verification[
                        "email"
                    ]
                )

                st.session_state.show_existing_dialog = True
                st.rerun()

            except Exception:

                st.error(
                    "We couldn't send the verification code right now. "
                    "Please wait a moment and try again."
                )

        st.divider()

        if st.button(
            "Forgot Household ID?",
            use_container_width=True,
        ):

            clear_verification_state()
            st.session_state.show_recovery_dialog = True
            st.rerun()

        return

    household_reference = (
        st.session_state.verification_reference
    )

    email = (
        st.session_state.verification_email
    )

    st.success(
        "Check your email."
    )

    st.write(
        "We sent a 6-digit verification code to:"
    )

    st.write(
        f"**{mask_email(email)}**"
    )

    st.caption(
        "The code expires in 10 minutes."
    )

    verification_code = (
        st.text_input(
            "Verification Code",
            max_chars=6,
            placeholder="123456",
        )
    )

    if st.button(
        "Verify & Continue",
        type="primary",
        use_container_width=True,
    ):

        if not verification_code.strip():

            st.error(
                "Please enter the 6-digit verification code."
            )
            return

        verified, status = (
            verify_household_code(
                household_reference,
                verification_code,
            )
        )

        if not verified:

            if status == "expired":

                st.error(
                    "That verification code has expired. "
                    "Send yourself a new code to continue."
                )

            elif status == "locked":

                st.error(
                    "Too many incorrect attempts were made with "
                    "that code. Please request a new one."
                )

            elif status == "no_active_code":

                st.error(
                    "That verification code is no longer active. "
                    "Please request a new one."
                )

            else:

                st.error(
                    "That code doesn't match. "
                    "Please check the email and try again."
                )

            return

        result = (
            get_registration_by_reference(
                household_reference
            )
        )

        if result is None:

            st.error(
                "We verified your code, but couldn't load the "
                "registration right now. Please try again."
            )
            return

        household, children = (
            result
        )

        st.session_state.household = {
            "parent_a_first_name":
                household[
                    "parent_a_first_name"
                ],

            "parent_a_last_name":
                household[
                    "parent_a_last_name"
                ],

            "parent_a_email":
                household[
                    "parent_a_email"
                ],

            "parent_a_phone":
                household[
                    "parent_a_phone"
                ],

            "parent_b_first_name":
                household[
                    "parent_b_first_name"
                ]
                or "",

            "parent_b_last_name":
                household[
                    "parent_b_last_name"
                ]
                or "",

            "parent_b_email":
                household[
                    "parent_b_email"
                ]
                or "",

            "parent_b_phone":
                household[
                    "parent_b_phone"
                ]
                or "",

            "address_line_1":
                household[
                    "address_line_1"
                ],

            "address_line_2":
                household[
                    "address_line_2"
                ]
                or "",

            "city":
                household[
                    "city"
                ],

            "state":
                household[
                    "state"
                ],

            "zip_code":
                household[
                    "zip_code"
                ],

            "emergency_contact_name":
                household.get(
                    "emergency_contact_name"
                )
                or "",

            "emergency_contact_relationship":
                household.get(
                    "emergency_contact_relationship"
                )
                or "",

            "emergency_contact_phone":
                household.get(
                    "emergency_contact_phone"
                )
                or "",
        }

        st.session_state.children = children

        st.session_state.existing_household_id = (
            household[
                "household_id"
            ]
        )

        st.session_state.existing_household_reference = (
            household[
                "household_reference"
            ]
        )

        st.session_state.registration_mode = (
            "existing"
        )

        clear_verification_state()
        clear_recovery_state()
        clear_admin_login_state()

        st.rerun()

    st.divider()

    if st.button(
        "Send a New Code",
        use_container_width=True,
    ):

        try:

            verification = (
                create_household_verification(
                    household_reference
                )
            )

            if verification is None:

                st.error(
                    "We couldn't find that household registration. "
                    "Please return and check the Household ID."
                )
                return

            send_verification_email(
                recipient=verification[
                    "email"
                ],
                verification_code=verification[
                    "code"
                ],
                household_reference=verification[
                    "household_reference"
                ],
                expires_minutes=verification[
                    "expires_minutes"
                ],
            )

            st.session_state.verification_email = (
                verification[
                    "email"
                ]
            )

            st.success(
                "A new verification code has been sent."
            )

        except Exception:

            st.error(
                "We couldn't send a new verification code right now. "
                "Please wait a moment and try again."
            )

    if st.button(
        "Use a Different Household ID",
        use_container_width=True,
    ):

        st.session_state.verification_reference = None
        st.session_state.verification_email = None
        st.session_state.show_existing_dialog = True
        st.rerun()


# ---------------------------------------------------------
# Household dialog
# ---------------------------------------------------------

@st.dialog("Household Information")
def household_dialog():

    household = (
        st.session_state.household
        or {}
    )

    with st.form(
        "household_form",
        enter_to_submit=False,
    ):

        # -------------------------------------------------
        # Parent / Guardian A
        # -------------------------------------------------

        st.subheader(
            "Parent / Guardian A"
        )

        st.caption(
            "Required"
        )

        col1, col2 = st.columns(2)

        with col1:

            parent_a_first_name = (
                st.text_input(
                    "First name",
                    value=household.get(
                        "parent_a_first_name",
                        "",
                    ),
                )
            )

        with col2:

            parent_a_last_name = (
                st.text_input(
                    "Last name",
                    value=household.get(
                        "parent_a_last_name",
                        "",
                    ),
                )
            )

        col1, col2 = st.columns(2)

        with col1:

            parent_a_email = (
                st.text_input(
                    "Email",
                    value=household.get(
                        "parent_a_email",
                        "",
                    ),
                )
            )

        with col2:

            parent_a_phone = (
                st.text_input(
                    "Phone",
                    value=household.get(
                        "parent_a_phone",
                        "",
                    ),
                )
            )

        st.divider()

        # -------------------------------------------------
        # Parent / Guardian B
        # -------------------------------------------------

        st.subheader(
            "Parent / Guardian B"
        )

        st.caption(
            "Optional"
        )

        col1, col2 = st.columns(2)

        with col1:

            parent_b_first_name = (
                st.text_input(
                    "First name",
                    value=household.get(
                        "parent_b_first_name",
                        "",
                    ),
                    key="parent_b_first_name",
                )
            )

        with col2:

            parent_b_last_name = (
                st.text_input(
                    "Last name",
                    value=household.get(
                        "parent_b_last_name",
                        "",
                    ),
                    key="parent_b_last_name",
                )
            )

        col1, col2 = st.columns(2)

        with col1:

            parent_b_email = (
                st.text_input(
                    "Email",
                    value=household.get(
                        "parent_b_email",
                        "",
                    ),
                    key="parent_b_email",
                )
            )

        with col2:

            parent_b_phone = (
                st.text_input(
                    "Phone",
                    value=household.get(
                        "parent_b_phone",
                        "",
                    ),
                    key="parent_b_phone",
                )
            )

        st.divider()

        # -------------------------------------------------
        # Home Address
        # -------------------------------------------------

        st.subheader(
            "Home Address"
        )

        address_line_1 = (
            st.text_input(
                "Street address",
                value=household.get(
                    "address_line_1",
                    "",
                ),
            )
        )

        address_line_2 = (
            st.text_input(
                "Apartment, unit, etc.",
                value=household.get(
                    "address_line_2",
                    "",
                ),
            )
        )

        (
            city_col,
            state_col,
            zip_col,
        ) = st.columns(
            [
                2,
                1,
                1,
            ]
        )

        with city_col:

            city = st.text_input(
                "City",
                value=household.get(
                    "city",
                    "",
                ),
            )

        with state_col:

            state = st.text_input(
                "State",
                value=household.get(
                    "state",
                    "WV",
                ),
                max_chars=2,
            )

        with zip_col:

            zip_code = (
                st.text_input(
                    "ZIP",
                    value=household.get(
                        "zip_code",
                        "",
                    ),
                )
            )

        st.divider()

        # -------------------------------------------------
        # Emergency Contact
        # -------------------------------------------------

        st.subheader(
            "Emergency Contact"
        )

        st.caption(
            "Please provide someone other than the parents "
            "or guardians listed above who we may contact "
            "if we're unable to reach you."
        )

        emergency_contact_name = (
            st.text_input(
                "Emergency contact name",
                value=household.get(
                    "emergency_contact_name",
                    "",
                ),
            )
        )

        (
            emergency_relationship_col,
            emergency_phone_col,
        ) = st.columns(2)

        with emergency_relationship_col:

            emergency_contact_relationship = (
                st.text_input(
                    "Relationship",
                    value=household.get(
                        "emergency_contact_relationship",
                        "",
                    ),
                    placeholder="Grandparent, aunt, family friend...",
                )
            )

        with emergency_phone_col:

            emergency_contact_phone = (
                st.text_input(
                    "Emergency contact phone",
                    value=household.get(
                        "emergency_contact_phone",
                        "",
                    ),
                )
            )

        submitted = (
            st.form_submit_button(
                "Save Household",
                type="primary",
                use_container_width=True,
            )
        )

        if submitted:

            required_fields = {
                "Parent / Guardian A first name":
                    parent_a_first_name,

                "Parent / Guardian A last name":
                    parent_a_last_name,

                "Email":
                    parent_a_email,

                "Phone":
                    parent_a_phone,

                "Street address":
                    address_line_1,

                "City":
                    city,

                "State":
                    state,

                "ZIP":
                    zip_code,

                "Emergency contact name":
                    emergency_contact_name,

                "Emergency contact relationship":
                    emergency_contact_relationship,

                "Emergency contact phone":
                    emergency_contact_phone,
            }

            missing = [
                name

                for name, value
                in required_fields.items()

                if not value.strip()
            ]

            if missing:

                st.error(
                    "Please complete: "
                    + ", ".join(
                        missing
                    )
                    + "."
                )
                return

            if not is_valid_email(
                parent_a_email
            ):

                st.error(
                    "Please enter a valid email address "
                    "for Parent / Guardian A."
                )
                return

            if (
                parent_b_email.strip()
                and not is_valid_email(
                    parent_b_email
                )
            ):

                st.error(
                    "Please enter a valid email address "
                    "for Parent / Guardian B."
                )
                return

            parent_a_phone_formatted = (
                normalize_phone(
                    parent_a_phone
                )
            )

            if (
                parent_a_phone_formatted
                is None
            ):

                st.error(
                    "Please enter a valid 10-digit phone number "
                    "for Parent / Guardian A."
                )
                return

            parent_b_phone_formatted = None

            if parent_b_phone.strip():

                parent_b_phone_formatted = (
                    normalize_phone(
                        parent_b_phone
                    )
                )

                if (
                    parent_b_phone_formatted
                    is None
                ):

                    st.error(
                        "Please enter a valid 10-digit phone number "
                        "for Parent / Guardian B."
                    )
                    return

            emergency_phone_formatted = (
                normalize_phone(
                    emergency_contact_phone
                )
            )

            if (
                emergency_phone_formatted
                is None
            ):

                st.error(
                    "Please enter a valid 10-digit phone number "
                    "for the emergency contact."
                )
                return

            zip_code_formatted = (
                normalize_zip(
                    zip_code
                )
            )

            if (
                zip_code_formatted
                is None
            ):

                st.error(
                    "Please enter a valid ZIP code, "
                    "such as 25033 or 25033-1234."
                )
                return

            state_formatted = (
                state
                .strip()
                .upper()
            )

            if not re.fullmatch(
                r"[A-Z]{2}",
                state_formatted,
            ):

                st.error(
                    "Please enter the two-letter state abbreviation, "
                    "such as WV."
                )
                return

            st.session_state.household = {
                "parent_a_first_name":
                    parent_a_first_name.strip(),

                "parent_a_last_name":
                    parent_a_last_name.strip(),

                "parent_a_email":
                    parent_a_email.strip(),

                "parent_a_phone":
                    parent_a_phone_formatted,

                "parent_b_first_name":
                    parent_b_first_name.strip(),

                "parent_b_last_name":
                    parent_b_last_name.strip(),

                "parent_b_email":
                    parent_b_email.strip(),

                "parent_b_phone":
                    parent_b_phone_formatted
                    or "",

                "address_line_1":
                    address_line_1.strip(),

                "address_line_2":
                    address_line_2.strip(),

                "city":
                    city.strip(),

                "state":
                    state_formatted,

                "zip_code":
                    zip_code_formatted,

                "emergency_contact_name":
                    emergency_contact_name.strip(),

                "emergency_contact_relationship":
                    emergency_contact_relationship.strip(),

                "emergency_contact_phone":
                    emergency_phone_formatted,
            }

            st.rerun()


# ---------------------------------------------------------
# Child dialog
# ---------------------------------------------------------

@st.dialog("Child Information")
def child_dialog(
    child_index: int | None = None,
):

    editing = (
        child_index
        is not None
    )

    if editing:

        child = (
            st.session_state.children[
                child_index
            ]
        )

    else:

        child = {}

    household = (
        st.session_state.household
        or {}
    )

    default_last_name = (
        household.get(
            "parent_a_last_name",
            "",
        )
    )

    grades = [
        "Select grade",
        "Pre-K",
        "K",
        "1",
        "2",
        "3",
        "4",
        "5",
        "6",
        "7",
        "8",
        "9",
        "10",
        "11",
        "12",
    ]

    existing_grade = (
        child.get(
            "grade",
            "Select grade",
        )
    )

    try:

        grade_index = (
            grades.index(
                existing_grade
            )
        )

    except ValueError:

        grade_index = 0

    sacrament_status_options = [
        "Select one",
        "Yes",
        "No",
        "Not sure",
    ]

    st.subheader(
        "Basic Information"
    )

    col1, col2 = st.columns(2)

    with col1:

        first_name = (
            st.text_input(
                "First name",
                value=child.get(
                    "first_name",
                    "",
                ),
            )
        )

    with col2:

        middle_name = (
            st.text_input(
                "Middle name",
                value=child.get(
                    "middle_name",
                    "",
                ),
            )
        )

    last_name = (
        st.text_input(
            "Last name",
            value=child.get(
                "last_name",
                default_last_name,
            ),
        )
    )

    today = date.today()

    earliest_birth_date = date(
        today.year - 20,
        1,
        1,
    )

    date_of_birth = (
        st.date_input(
            "Date of birth",
            value=child.get(
                "date_of_birth",
                None,
            ),
            min_value=earliest_birth_date,
            max_value=today,
        )
    )

    grade = st.selectbox(
        "Grade",
        grades,
        index=grade_index,
    )

    school = st.text_input(
        "School",
        value=child.get(
            "school",
            "",
        ),
    )

    st.divider()

    st.subheader(
        "Sacrament Preparation"
    )

    st.caption(
        "Select any sacraments this child is preparing "
        "to receive this year."
    )

    receiving_first_communion_reconciliation = (
        st.toggle(
            "Receiving First Reconciliation / "
            "First Communion this year",
            value=child.get(
                "receiving_first_communion_reconciliation",
                False,
            ),
        )
    )

    receiving_confirmation = (
        st.toggle(
            "Receiving Confirmation this year",
            value=child.get(
                "receiving_confirmation",
                False,
            ),
        )
    )

    baptism_status = None
    first_reconciliation_status = None
    first_communion_status = None

    sacramental_history_needed = (
        receiving_first_communion_reconciliation
        or receiving_confirmation
    )

    if sacramental_history_needed:

        st.divider()

        st.subheader(
            "Sacramental History"
        )

        st.caption(
            "Please tell us which sacraments this child "
            "has already received."
        )

        baptism_status = (
            st.selectbox(
                "Has this child been baptized?",
                sacrament_status_options,
                index=sacrament_status_index(
                    child.get(
                        "baptism_status"
                    )
                ),
            )
        )

        if receiving_confirmation:

            first_reconciliation_status = (
                st.selectbox(
                    "Has this child received "
                    "First Reconciliation?",
                    sacrament_status_options,
                    index=sacrament_status_index(
                        child.get(
                            "first_reconciliation_status"
                        )
                    ),
                )
            )

            first_communion_status = (
                st.selectbox(
                    "Has this child received "
                    "First Communion?",
                    sacrament_status_options,
                    index=sacrament_status_index(
                        child.get(
                            "first_communion_status"
                        )
                    ),
                )
            )

        preview_child = {
            "receiving_first_communion_reconciliation":
                receiving_first_communion_reconciliation,

            "receiving_confirmation":
                receiving_confirmation,

            "baptism_status":
                (
                    None
                    if baptism_status
                    == "Select one"
                    else baptism_status
                ),

            "first_reconciliation_status":
                (
                    None
                    if first_reconciliation_status
                    in (
                        None,
                        "Select one",
                    )
                    else first_reconciliation_status
                ),

            "first_communion_status":
                (
                    None
                    if first_communion_status
                    in (
                        None,
                        "Select one",
                    )
                    else first_communion_status
                ),
        }

        follow_up_reasons = (
            sacramental_follow_up_reasons(
                preview_child
            )
        )

        history_questions_complete = (
            baptism_status
            != "Select one"
        )

        if receiving_confirmation:

            history_questions_complete = (
                history_questions_complete
                and first_reconciliation_status
                != "Select one"
                and first_communion_status
                != "Select one"
            )

        if (
            history_questions_complete
            and follow_up_reasons
        ):

            st.warning(
                "This registration will need sacramental follow-up. "
                "You can still continue normally. A member of the "
                "faith formation team will contact you if needed."
            )

    st.divider()

    button_text = (
        "Save Changes"
        if editing
        else "Add Child"
    )

    if st.button(
        button_text,
        type="primary",
        use_container_width=True,
    ):

        if not first_name.strip():

            st.error(
                "Please enter the child's first name."
            )
            return

        if not last_name.strip():

            st.error(
                "Please enter the child's last name."
            )
            return

        if not school.strip():

            st.error(
                "Please enter the child's school."
            )
            return

        if date_of_birth is None:

            st.error(
                "Please enter the child's date of birth."
            )
            return

        if grade == "Select grade":

            st.error(
                "Please select the child's grade."
            )
            return

        if sacramental_history_needed:

            if (
                baptism_status
                == "Select one"
            ):

                st.error(
                    "Please tell us whether this child has been baptized."
                )
                return

        if receiving_confirmation:

            if (
                first_reconciliation_status
                == "Select one"
            ):

                st.error(
                    "Please tell us whether this child has received "
                    "First Reconciliation."
                )
                return

            if (
                first_communion_status
                == "Select one"
            ):

                st.error(
                    "Please tell us whether this child has received "
                    "First Communion."
                )
                return

        if sacramental_history_needed:

            saved_baptism_status = (
                baptism_status
            )

        else:

            saved_baptism_status = (
                child.get(
                    "baptism_status"
                )
            )

        if receiving_confirmation:

            saved_first_reconciliation_status = (
                first_reconciliation_status
            )

            saved_first_communion_status = (
                first_communion_status
            )

        else:

            saved_first_reconciliation_status = (
                child.get(
                    "first_reconciliation_status"
                )
            )

            saved_first_communion_status = (
                child.get(
                    "first_communion_status"
                )
            )

        child_data = {
            "first_name":
                first_name.strip(),

            "middle_name":
                middle_name.strip(),

            "last_name":
                last_name.strip(),

            "date_of_birth":
                date_of_birth,

            "grade":
                grade,

            "school":
                school.strip(),

            "receiving_first_communion_reconciliation":
                receiving_first_communion_reconciliation,

            "receiving_confirmation":
                receiving_confirmation,

            "baptism_status":
                saved_baptism_status,

            "first_reconciliation_status":
                saved_first_reconciliation_status,

            "first_communion_status":
                saved_first_communion_status,
        }

        if (
            editing
            and child.get(
                "child_id"
            )
            is not None
        ):

            child_data[
                "child_id"
            ] = child[
                "child_id"
            ]

        if editing:

            st.session_state.children[
                child_index
            ] = child_data

        else:

            st.session_state.children.append(
                child_data
            )

        st.rerun()


# ---------------------------------------------------------
# Review dialog
# ---------------------------------------------------------

@st.dialog(
    "Review Registration",
    width="medium",
)
def review_dialog():

    household = (
        st.session_state.household
    )

    children = (
        st.session_state.children
    )

    editing_existing = (
        st.session_state.registration_mode
        == "existing"
    )

    st.markdown(
        """
        <div class="review-section-title">
            Household
        </div>
        """,
        unsafe_allow_html=True,
    )

    parent_a = parent_name(
        household.get(
            "parent_a_first_name"
        ),
        household.get(
            "parent_a_last_name"
        ),
    )

    parent_b = parent_name(
        household.get(
            "parent_b_first_name"
        ),
        household.get(
            "parent_b_last_name"
        ),
    )

    with st.container(
        border=True,
        key="review_dialog_household_card",
    ):

        st.markdown(
            f"""
            <div class="review-field-label">
                Parent / Guardian A
            </div>

            <div class="review-person-name">
                {escape_html(parent_a)}
            </div>

            <div class="review-field-label">
                Email
            </div>

            <div class="review-field-value">
                {escape_html(household['parent_a_email'])}
            </div>

            <div class="review-field-label">
                Phone
            </div>

            <div class="review-field-value">
                {escape_html(household['parent_a_phone'])}
            </div>
            """,
            unsafe_allow_html=True,
        )

        if (
            parent_b
            or household.get(
                "parent_b_email"
            )
            or household.get(
                "parent_b_phone"
            )
        ):

            st.divider()

            st.markdown(
                """
                <div class="review-field-label">
                    Parent / Guardian B
                </div>
                """,
                unsafe_allow_html=True,
            )

            if parent_b:

                st.markdown(
                    f"""
                    <div class="review-person-name">
                        {escape_html(parent_b)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            if household.get(
                "parent_b_email"
            ):

                st.markdown(
                    f"""
                    <div class="review-field-label">
                        Email
                    </div>

                    <div class="review-field-value">
                        {escape_html(household['parent_b_email'])}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            if household.get(
                "parent_b_phone"
            ):

                st.markdown(
                    f"""
                    <div class="review-field-label">
                        Phone
                    </div>

                    <div class="review-field-value">
                        {escape_html(household['parent_b_phone'])}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        st.divider()

        address_html = (
            escape_html(
                household[
                    "address_line_1"
                ]
            )
        )

        if household.get(
            "address_line_2"
        ):

            address_html += (
                "<br>"
                + escape_html(
                    household[
                        "address_line_2"
                    ]
                )
            )

        address_html += (
            "<br>"
            f"{escape_html(household['city'])}, "
            f"{escape_html(household['state'])} "
            f"{escape_html(household['zip_code'])}"
        )

        st.markdown(
            f"""
            <div class="review-field-label">
                Home Address
            </div>

            <div class="review-field-value">
                {address_html}
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.divider()

        st.markdown(
            f"""
            <div class="review-field-label">
                Emergency Contact
            </div>

            <div class="review-person-name">
                {escape_html(household['emergency_contact_name'])}
            </div>

            <div class="review-field-label">
                Relationship
            </div>

            <div class="review-field-value">
                {escape_html(household['emergency_contact_relationship'])}
            </div>

            <div class="review-field-label">
                Phone
            </div>

            <div class="review-field-value">
                {escape_html(household['emergency_contact_phone'])}
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.write("")

    st.markdown(
        """
        <div class="review-section-title">
            Children
        </div>
        """,
        unsafe_allow_html=True,
    )

    for child_index, child in enumerate(children):

        child_name = full_name(
            child.get(
                "first_name"
            ),
            child.get(
                "middle_name"
            ),
            child.get(
                "last_name"
            ),
        )

        age = calculate_age(
            child[
                "date_of_birth"
            ]
        )

        with st.container(
            border=True,
            key=f"review_dialog_child_card_{child_index}",
        ):

            st.markdown(
                f"""
                <div class="review-child-name">
                    {escape_html(child_name)}
                </div>

                <div class="review-field-label">
                    Grade
                </div>

                <div class="review-field-value">
                    {escape_html(child_grade_label(child['grade']))}
                </div>

                <div class="review-field-label">
                    School
                </div>

                <div class="review-field-value">
                    {escape_html(child['school'])}
                </div>

                <div class="review-field-label">
                    Date of Birth
                </div>

                <div class="review-field-value">
                    {child['date_of_birth'].strftime('%m/%d/%Y')}
                </div>

                <div class="review-field-label">
                    Age
                </div>

                <div class="review-field-value">
                    {age}
                </div>
                """,
                unsafe_allow_html=True,
            )

            preparation = (
                sacrament_preparation_labels(
                    child
                )
            )

            if preparation:

                preparation_text = (
                    ", ".join(
                        preparation
                    )
                )

                st.markdown(
                    f"""
                    <div class="review-field-label">
                        Sacrament Preparation
                    </div>

                    <div class="review-field-value">
                        {escape_html(preparation_text)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            history_parts = []

            if child.get(
                "baptism_status"
            ):

                history_parts.append(
                    (
                        "Baptism",
                        child[
                            "baptism_status"
                        ],
                    )
                )

            if child.get(
                "first_reconciliation_status"
            ):

                history_parts.append(
                    (
                        "First Reconciliation",
                        child[
                            "first_reconciliation_status"
                        ],
                    )
                )

            if child.get(
                "first_communion_status"
            ):

                history_parts.append(
                    (
                        "First Communion",
                        child[
                            "first_communion_status"
                        ],
                    )
                )

            if history_parts:

                for (
                    sacrament_name,
                    sacrament_status,
                ) in history_parts:

                    st.markdown(
                        f"""
                        <div class="review-field-label">
                            {escape_html(sacrament_name)}
                        </div>

                        <div class="review-field-value">
                            {escape_html(sacrament_status)}
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

            follow_up_reasons = (
                sacramental_follow_up_reasons(
                    child
                )
            )

            if follow_up_reasons:

                st.warning(
                    "Sacramental follow-up will be needed for this child. "
                    "You can still submit the registration normally."
                )

    st.write("")

    with st.container(
        border=True,
        key="review_dialog_submit_card",
    ):

        st.markdown(
            """
            <div class="review-submit-title">
                Take a moment to review your information.
            </div>

            <div class="review-submit-copy">
                If everything above looks correct, submit
                your registration below.
            </div>

            <div style="height: 0.45rem;"></div>
            """,
            unsafe_allow_html=True,
        )

    st.write("")

    submit_label = (
        "Save Changes"
        if editing_existing
        else "Submit Registration"
    )

    if st.button(
        submit_label,
        type="primary",
        use_container_width=True,
    ):

        st.session_state.confirmation_email_sent = None
        st.session_state.confirmation_email_address = None
        st.session_state.confirmation_email_error = None

        recipient_email = (
            household[
                "parent_a_email"
            ]
        )

        try:

            if editing_existing:

                update_registration(
                    st.session_state.existing_household_id,
                    household,
                    children,
                )

                household_id = (
                    st.session_state.existing_household_id
                )

                household_reference = (
                    st.session_state.existing_household_reference
                )

                st.session_state.submitted_household_id = (
                    household_id
                )

                st.session_state.submitted_household_reference = (
                    household_reference
                )

                try:

                    send_update_confirmation(
                        recipient=recipient_email,
                        household_reference=(
                            household_reference
                        ),
                    )

                    st.session_state.confirmation_email_sent = True

                    st.session_state.confirmation_email_address = (
                        recipient_email
                    )

                except Exception:

                    st.session_state.confirmation_email_sent = False

                    st.session_state.confirmation_email_address = (
                        recipient_email
                    )

                    st.session_state.confirmation_email_error = (
                        "Confirmation email could not be sent."
                    )

            else:

                (
                    household_id,
                    household_reference,
                ) = save_registration(
                    household,
                    children,
                )

                st.session_state.submitted_household_id = (
                    household_id
                )

                st.session_state.submitted_household_reference = (
                    household_reference
                )

                try:

                    send_registration_confirmation(
                        recipient=recipient_email,
                        household_reference=(
                            household_reference
                        ),
                        children=children,
                    )

                    st.session_state.confirmation_email_sent = True

                    st.session_state.confirmation_email_address = (
                        recipient_email
                    )

                except Exception:

                    st.session_state.confirmation_email_sent = False

                    st.session_state.confirmation_email_address = (
                        recipient_email
                    )

                    st.session_state.confirmation_email_error = (
                        "Confirmation email could not be sent."
                    )

            st.session_state.household = None
            st.session_state.children = []

            st.rerun()

        except Exception:

            st.error(
                "We couldn't save your registration right now. "
                "Your information is still here, so you won't need "
                "to enter it again. Please wait a moment and try again."
            )


# ---------------------------------------------------------
# Roster card
# ---------------------------------------------------------

def render_roster_card(
    group: dict,
    roster: list[dict],
) -> None:

    group_key = (
        group[
            "group_key"
        ]
    )

    title = roster_title(
        group_key,
        group[
            "display_name"
        ],
    )

    group_children = [
        child

        for child in roster

        if (
            roster_group_key_for_grade(
                child[
                    "grade"
                ]
            )
            == group_key
        )
    ]

    with st.container(
        border=True,
        key=f"roster_card_{group_key}",
    ):

        # -------------------------------------------------
        # Roster title / Catechist edit
        # -------------------------------------------------

        title_col, edit_col = (
            st.columns(
                [
                    5,
                    1.2,
                ],
                vertical_alignment="center",
            )
        )

        with title_col:

            st.subheader(
                f"{title}  ·  "
                f"{len(group_children)}"
            )

        with edit_col:

            if st.button(
                "Edit",
                key=(
                    f"edit_catechists_"
                    f"{group_key}"
                ),
                use_container_width=True,
            ):

                edit_catechists_dialog(
                    group
                )

        # -------------------------------------------------
        # Catechists
        # -------------------------------------------------

        catechists = (
            group.get(
                "catechists",
                "",
            )
            .strip()
        )

        if catechists:

            st.write(
                f"**Catechists:** "
                f"{catechists}"
            )

        else:

            st.caption(
                "Catechists: Not assigned"
            )

        # -------------------------------------------------
        # Classroom
        # -------------------------------------------------

        classroom = (
            group.get(
                "classroom",
                "",
            )
            or ""
        )

        classroom_value = (
            st.text_input(
                "Classroom",
                value=classroom,
                placeholder="Room 1 - St. Monica",
                key=(
                    f"classroom_"
                    f"{group_key}"
                ),
            )
        )

        if (
            classroom_value.strip()
            != classroom.strip()
        ):

            if st.button(
                "Save Classroom",
                key=(
                    f"save_classroom_"
                    f"{group_key}"
                ),
                use_container_width=True,
            ):

                try:

                    update_roster_group_classroom(
                        group_key,
                        classroom_value,
                    )

                    st.rerun()

                except Exception:

                    st.error(
                        "We couldn't save the classroom right now. "
                        "Please try again."
                    )

        st.write("")

        # -------------------------------------------------
        # Children
        # -------------------------------------------------

        if not group_children:

            st.info(
                "No children are currently "
                "registered in this roster."
            )
            return

        rows = []

        for child in group_children:

            child_name = full_name(
                child.get(
                    "first_name"
                ),
                child.get(
                    "middle_name"
                ),
                child.get(
                    "last_name"
                ),
            )

            if group_key in (
                "kindergarten",
                "edge",
                "life_teen",
            ):

                rows.append(
                    {
                        "Child":
                            child_name,

                        "Grade":
                            child[
                                "grade"
                            ],

                        "School":
                            child[
                                "school"
                            ],
                    }
                )

            else:

                rows.append(
                    {
                        "Child":
                            child_name,

                        "School":
                            child[
                                "school"
                            ],
                    }
                )

        display_df = (
            pd.DataFrame(
                rows
            )
        )

        st.caption(
            "Select a child to view details."
        )

        table_event = (
            st.dataframe(
                display_df,
                hide_index=True,
                use_container_width=True,
                on_select="rerun",
                selection_mode="single-row",
                key=(
                    f"roster_table_"
                    f"{group_key}_"
                    f"{st.session_state.admin_detail_table_nonce}"
                ),
            )
        )

        selected_rows = (
            table_event
            .selection
            .rows
        )

        if selected_rows:

            selected_index = (
                selected_rows[0]
            )

            if (
                0
                <= selected_index
                < len(
                    group_children
                )
            ):

                selected_child = (
                    group_children[
                        selected_index
                    ]
                )

                st.session_state.admin_detail_child_id = (
                    selected_child[
                        "child_id"
                    ]
                )

        # -------------------------------------------------
        # Export
        # -------------------------------------------------

        pdf_data = (
            build_roster_pdf(
                title=title,
                classroom=(
                    classroom_value
                    .strip()
                ),
                catechists=catechists,
                children=group_children,
            )
        )

        file_group_name = (
            group_key.replace(
                "_",
                "-"
            )
        )

        st.download_button(
            f"Download {title} Roster",
            data=pdf_data,
            file_name=(
                f"ascension-"
                f"{file_group_name}-"
                f"roster.pdf"
            ),
            mime="application/pdf",
            key=(
                f"download_roster_"
                f"{group_key}"
            ),
            use_container_width=True,
            help=(
                "Download a print-ready PDF roster."
            ),
        )


        attendance_data = (
            build_attendance_sheet(
                title=title,
                classroom=(
                    classroom_value
                    .strip()
                ),
                catechists=catechists,
                children=group_children,
            )
        )

        st.download_button(
            f"Download {title} Attendance Sheet",
            data=attendance_data,
            file_name=(
                f"ascension-"
                f"{file_group_name}-"
                f"attendance-"
                f"{date.today().isoformat()}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            key=(
                f"download_attendance_"
                f"{group_key}"
            ),
            use_container_width=True,
            help=(
                "Download an Excel attendance sheet for the next "
                "8 Sundays."
            ),
        )


# ---------------------------------------------------------
# Admin dashboard
# ---------------------------------------------------------

if (
    st.session_state.admin_authenticated
):

    if not is_authorized_admin(
        st.session_state.admin_email
        or ""
    ):

        st.session_state.admin_authenticated = False
        st.session_state.admin_email = None

        clear_admin_login_state()

        st.rerun()

    title_col, refresh_col, logout_col = (
        st.columns(
            [
                4.5,
                1.25,
                1.25,
            ],
            vertical_alignment="center",
        )
    )

    with title_col:

        st.title(
            "Ascension Registration"
        )

        st.caption(
            f"Admin Dashboard • "
            f"{st.session_state.admin_email}"
        )

    with refresh_col:

        if st.button(
            "↻ Refresh",
            use_container_width=True,
            help="Reload registration data without signing out.",
        ):

            # Rerun the Streamlit script while preserving the
            # current Session State, including admin authentication.
            st.rerun()

    with logout_col:

        if st.button(
            "Log Out",
            use_container_width=True,
        ):

            st.session_state.admin_authenticated = False
            st.session_state.admin_email = None

            clear_admin_login_state()
            clear_admin_child_detail()

            st.rerun()

    try:

        roster = (
            get_admin_roster()
        )

        roster_groups = (
            get_roster_groups()
        )

    except Exception:

        st.error(
            "We couldn't load the registration dashboard right now. "
            "Please refresh the page or try again in a moment."
        )

        st.stop()

    st.header(
        "Registration Overview"
    )

    total_children = (
        len(
            roster
        )
    )

    total_households = len(
        {
            child[
                "household_reference"
            ]

            for child
            in roster
        }
    )

    first_communion_count = sum(
        1

        for child
        in roster

        if child.get(
            "receiving_first_communion_reconciliation",
            False,
        )
    )

    confirmation_count = sum(
        1

        for child
        in roster

        if child.get(
            "receiving_confirmation",
            False,
        )
    )

    follow_up_count = sum(
        1

        for child
        in roster

        if sacramental_follow_up_reasons(
            child
        )
    )

    (
        metric_1,
        metric_2,
        metric_3,
        metric_4,
        metric_5,
    ) = st.columns(5)

    with metric_1:

        st.metric(
            "Children",
            total_children,
        )

    with metric_2:

        st.metric(
            "Households",
            total_households,
        )

    with metric_3:

        st.metric(
            "First Communion",
            first_communion_count,
        )

    with metric_4:

        st.metric(
            "Confirmation",
            confirmation_count,
        )

    with metric_5:

        st.metric(
            "Follow-up",
            follow_up_count,
        )

    st.divider()

    st.header(
        "PSR Rosters"
    )

    psr_groups = [
        group

        for group
        in roster_groups

        if group[
            "category"
        ] == "PSR"
    ]

    for group in psr_groups:

        render_roster_card(
            group,
            roster,
        )

    st.divider()

    st.header(
        "Youth Ministry Rosters"
    )

    youth_groups = [
        group

        for group
        in roster_groups

        if group[
            "category"
        ]
        == "Youth Ministry"
    ]

    for group in youth_groups:

        render_roster_card(
            group,
            roster,
        )

    # -----------------------------------------------------
    # Upcoming birthdays
    # -----------------------------------------------------

    birthday_df = (
        get_upcoming_birthdays(
            roster
        )
    )

    today = date.today()

    next_month_date = (
        date(
            today.year + 1,
            1,
            1,
        )
        if today.month == 12
        else date(
            today.year,
            today.month + 1,
            1,
        )
    )

    birthday_month_label = (
        f"{today.strftime('%B')} & "
        f"{next_month_date.strftime('%B')}"
    )

    with st.expander(
        (
            "🎂 Upcoming Birthdays · "
            f"{birthday_month_label} "
            f"({len(birthday_df)})"
        ),
        expanded=False,
    ):

        st.caption(
            "Grouped PSR • EDGE • Life Teen • Remaining birthdays "
            "this month and all birthdays next month."
        )

        if birthday_df.empty:

            st.info(
                "No upcoming birthdays in this period."
            )

        else:

            st.dataframe(
                birthday_df,
                hide_index=True,
                use_container_width=True,
            )

            birthday_csv = (
                birthday_df
                .to_csv(
                    index=False
                )
                .encode(
                    "utf-8-sig"
                )
            )

            st.download_button(
                "Download Birthday List",
                data=birthday_csv,
                file_name=(
                    "ascension-birthdays-"
                    f"{today.strftime('%Y-%m')}-"
                    f"{next_month_date.strftime('%Y-%m')}.csv"
                ),
                mime="text/csv",
                use_container_width=True,
                help=(
                    "Download this birthday list as a CSV file."
                ),
            )

    st.divider()

    # -----------------------------------------------------
    # Household messaging - DEV dry run only
    # -----------------------------------------------------

    with st.expander(
        "Household Messaging",
        expanded=False,
    ):

        st.caption(
            "Development preview only. Household drafts can be "
            "queued for the local Android gateway."
        )

        # -------------------------------------------------
        # DEV gateway one-number test
        # -------------------------------------------------

        with st.expander(
            "DEV Gateway Test - One Number",
            expanded=True,
        ):

            st.warning(
                "TEST ONLY. This panel never selects a parish household. "
                "It creates exactly one queued recipient using the phone "
                "number entered below."
            )

            test_name_col, test_phone_col = (
                st.columns(2)
            )

            with test_name_col:

                gateway_test_name = st.text_input(
                    "Test recipient name",
                    placeholder="Daniel / Courtney / Test Phone",
                    key="gateway_test_name",
                )

            with test_phone_col:

                gateway_test_phone = st.text_input(
                    "Test phone number",
                    placeholder="(304) 555-1234",
                    key="gateway_test_phone",
                )

            gateway_test_text = st.text_area(
                "Test message",
                value="Ascension Messenger DEV gateway test.",
                height=110,
                key="gateway_test_message_text",
            )

            st.caption(
                f"{len(gateway_test_text)} characters  •  "
                "One manually entered recipient only"
            )

            # A stable request key makes this queue action idempotent.
            # If Streamlit processes the same button action twice, the
            # database returns the existing Test ID instead of inserting
            # a second recipient.
            payload_signature = (
                gateway_test_name.strip(),
                gateway_test_phone.strip(),
                gateway_test_text.strip(),
            )

            if (
                st.session_state.get(
                    "gateway_test_payload_signature"
                ) != payload_signature
            ):
                st.session_state.gateway_test_payload_signature = (
                    payload_signature
                )
                st.session_state.gateway_test_request_key = (
                    uuid.uuid4().hex
                )
                st.session_state.pop(
                    "gateway_test_last_queued_id",
                    None,
                )

            if not st.session_state.get(
                "gateway_test_request_key"
            ):
                st.session_state.gateway_test_request_key = (
                    uuid.uuid4().hex
                )

            gateway_test_disabled = (
                not gateway_test_name.strip()
                or not gateway_test_phone.strip()
                or not gateway_test_text.strip()
            )

            if st.button(
                "Queue One-Number Test",
                type="primary",
                use_container_width=True,
                disabled=gateway_test_disabled,
                key="queue_gateway_test_message",
            ):

                try:

                    test_message_id = (
                        create_gateway_test_message(
                            created_by=(
                                st.session_state.admin_email
                                or ""
                            ),
                            contact_name=gateway_test_name,
                            phone=gateway_test_phone,
                            message_text=gateway_test_text,
                            request_key=(
                                st.session_state
                                .gateway_test_request_key
                            ),
                        )
                    )

                    st.session_state.gateway_test_last_queued_id = (
                        test_message_id
                    )

                    st.rerun()

                except Exception as exc:

                    st.error(
                        "We couldn't queue the gateway test."
                    )

                    st.caption(
                        f"DEV detail: {exc}"
                    )

            last_queued_test_id = st.session_state.get(
                "gateway_test_last_queued_id"
            )

            if last_queued_test_id:
                st.success(
                    f"TEST #{last_queued_test_id} is queued. Repeated "
                    "submissions of this same test request will reuse "
                    "that Test ID instead of creating a duplicate."
                )

                if st.button(
                    "Prepare Another Identical Test",
                    use_container_width=True,
                    key="prepare_another_identical_gateway_test",
                ):
                    st.session_state.gateway_test_request_key = (
                        uuid.uuid4().hex
                    )
                    st.session_state.pop(
                        "gateway_test_last_queued_id",
                        None,
                    )
                    st.rerun()

            st.write("")
            st.markdown("**Recent gateway tests**")

            try:

                gateway_tests = (
                    get_gateway_test_messages(
                        limit=20
                    )
                )

            except Exception as exc:

                gateway_tests = []

                st.error(
                    "We couldn't load DEV gateway tests."
                )

                st.caption(
                    f"DEV detail: {exc}"
                )

            if not gateway_tests:

                st.info(
                    "No one-number gateway tests have been queued yet."
                )

            else:

                gateway_test_rows = []

                for test_message in gateway_tests:

                    created_at = test_message.get(
                        "created_at"
                    )

                    created_label = (
                        created_at.strftime(
                            "%m/%d/%Y %I:%M %p"
                        )
                        if created_at is not None
                        else ""
                    )

                    gateway_test_rows.append(
                        {
                            "Test ID": test_message.get(
                                "message_id"
                            ),
                            "Created": created_label,
                            "Recipient": test_message.get(
                                "contact_name",
                                "",
                            ),
                            "Phone": test_message.get(
                                "phone",
                                "",
                            ),
                            "Status": str(
                                test_message.get(
                                    "recipient_status",
                                    "",
                                )
                            ).title(),
                            "Message": test_message.get(
                                "message_text",
                                "",
                            ),
                        }
                    )

                st.dataframe(
                    pd.DataFrame(
                        gateway_test_rows
                    ),
                    hide_index=True,
                    use_container_width=True,
                    column_config={
                        "Message":
                            st.column_config.TextColumn(
                                "Message",
                                width="large",
                            ),
                    },
                )

                cancellable_tests = [
                    test_message
                    for test_message in gateway_tests
                    if test_message.get(
                        "recipient_status"
                    ) == "queued"
                ]

                if cancellable_tests:

                    test_by_id = {
                        test_message[
                            "message_id"
                        ]:
                            test_message
                        for test_message
                        in cancellable_tests
                    }

                    selected_test_id = st.selectbox(
                        "Queued test to cancel / clear",
                        list(
                            test_by_id.keys()
                        ),
                        format_func=(
                            lambda message_id: (
                                f"TEST #{message_id} · "
                                + str(
                                    test_by_id[
                                        message_id
                                    ].get(
                                        "contact_name",
                                        "",
                                    )
                                )
                                + " · Queued"
                            )
                        ),
                        key="gateway_test_cancel_id",
                    )

                    if st.button(
                        "Cancel / Clear Selected Queued Test",
                        use_container_width=True,
                        key=(
                            "cancel_gateway_test_"
                            f"{selected_test_id}"
                        ),
                    ):

                        try:

                            cancelled = (
                                cancel_gateway_test_message(
                                    selected_test_id
                                )
                            )

                            if cancelled:

                                if (
                                    st.session_state.get(
                                        "gateway_test_last_queued_id"
                                    ) == selected_test_id
                                ):
                                    st.session_state.pop(
                                        "gateway_test_last_queued_id",
                                        None,
                                    )

                                st.success(
                                    f"TEST #{selected_test_id} cancelled. "
                                    "The gateway will not claim it."
                                )

                                st.rerun()

                            else:

                                st.warning(
                                    "That test is no longer safely "
                                    "cancellable from Streamlit. If the "
                                    "Pixel has already claimed it, cancel "
                                    "it from the Pixel instead."
                                )

                        except Exception as exc:

                            st.error(
                                "We couldn't cancel that gateway test."
                            )

                            st.caption(
                                f"DEV detail: {exc}"
                            )

                else:

                    st.caption(
                        "No queued tests are available to cancel here. "
                        "A claimed test must be cancelled by the Pixel "
                        "that owns its claim token."
                    )

            st.caption(
                "Gateway tests are isolated from parish household messages. "
                "The Pixel's DEV test fetch endpoint can claim only records "
                "tagged as tests."
            )

        st.divider()

        audiences = st.multiselect(
            "Audience",
            [
                "All Households",
                "PSR",
                "EDGE",
                "Life Teen",
                "Confirmation",
                "First Communion",
            ],
            default=[
                "EDGE",
                "Life Teen",
            ],
            placeholder="Select one or more groups",
            key="messaging_audiences",
        )

        if (
            "All Households" in audiences
            and len(audiences) > 1
        ):

            st.caption(
                "All Households is selected, so the other "
                "audience selections are included automatically."
            )

        (
            recipient_df,
            missing_phone_df,
            messaging_stats,
        ) = resolve_household_message_recipients(
            roster,
            audiences,
        )

        if not audiences:

            st.info(
                "Select at least one audience to build "
                "the household recipient list."
            )

        (
            message_metric_1,
            message_metric_2,
            message_metric_3,
            message_metric_4,
        ) = st.columns(4)

        with message_metric_1:

            st.metric(
                "Children",
                messaging_stats[
                    "matched_children"
                ],
            )

        with message_metric_2:

            st.metric(
                "Households",
                messaging_stats[
                    "unique_households"
                ],
            )

        with message_metric_3:

            st.metric(
                "Recipients",
                messaging_stats[
                    "recipients"
                ],
            )

        with message_metric_4:

            st.metric(
                "Missing Phone",
                messaging_stats[
                    "missing"
                ],
            )

        st.caption(
            f"Parent A used: {messaging_stats['parent_a']}  •  "
            f"Parent B fallback: {messaging_stats['parent_b']}"
        )

        st.write("")

        message_text = st.text_area(
            "Message",
            height=170,
            placeholder=(
                "Write the message you want to send to the "
                "selected households..."
            ),
            key="messaging_message_text",
        )

        st.caption(
            f"{len(message_text)} characters"
        )

        st.write("")

        st.subheader(
            "Recipient Preview"
        )

        if recipient_df.empty:

            st.info(
                "No households with a usable parent phone number "
                "match this audience."
            )

        else:

            st.dataframe(
                recipient_df,
                hide_index=True,
                use_container_width=True,
            )

            recipient_csv = (
                recipient_df
                .to_csv(
                    index=False
                )
                .encode(
                    "utf-8-sig"
                )
            )

            st.download_button(
                "Download Recipient List",
                data=recipient_csv,
                file_name=(
                    "ascension-household-message-recipients-"
                    + (
                        "all-households"
                        if "All Households" in audiences
                        else (
                            "-".join(
                                audience
                                .lower()
                                .replace(" ", "-")
                                for audience
                                in audiences
                            )
                            or "none"
                        )
                    )
                    + "-"
                    + date.today().isoformat()
                    + ".csv"
                ),
                mime="text/csv",
                use_container_width=True,
                key="download_message_recipients",
            )

        if not missing_phone_df.empty:

            with st.expander(
                (
                    "Households Missing a Phone Number "
                    f"({len(missing_phone_df)})"
                ),
                expanded=False,
            ):

                st.dataframe(
                    missing_phone_df,
                    hide_index=True,
                    use_container_width=True,
                )

        st.write("")

        message_action_disabled = (
            not audiences
            or recipient_df.empty
            or not message_text.strip()
        )

        preview_col, save_col = (
            st.columns(2)
        )

        with preview_col:

            if st.button(
                "Run Message Preview",
                use_container_width=True,
                disabled=message_action_disabled,
                key="messaging_dry_run",
            ):

                st.success(
                    f"Dry run complete. "
                    f"{len(recipient_df)} individual household "
                    f"{'message' if len(recipient_df) == 1 else 'messages'} "
                    "would be queued. No texts were sent."
                )

                st.write(
                    "**Message that would be sent:**"
                )

                st.code(
                    message_text,
                    language=None,
                )

        with save_col:

            if st.button(
                "Save Draft",
                type="primary",
                use_container_width=True,
                disabled=message_action_disabled,
                key="messaging_save_draft",
            ):

                try:

                    message_id = (
                        create_message_draft(
                            created_by=(
                                st.session_state.admin_email
                                or ""
                            ),
                            message_text=message_text,
                            audiences=audiences,
                            recipients=(
                                recipient_df
                                .to_dict(
                                    orient="records"
                                )
                            ),
                        )
                    )

                    st.success(
                        f"Draft #{message_id} saved with "
                        f"{len(recipient_df)} household "
                        f"{'recipient' if len(recipient_df) == 1 else 'recipients'}. "
                        "No texts were sent."
                    )

                except Exception as exc:

                    st.error(
                        "We couldn't save this message draft. "
                        "Please try again."
                    )

                    st.caption(
                        f"DEV detail: {exc}"
                    )

        st.divider()

        # -------------------------------------------------
        # Message history
        # -------------------------------------------------

        st.subheader(
            "Message History"
        )

        st.caption(
            "Saved drafts keep a snapshot of the exact household "
            "contacts and phone numbers selected at that time."
        )

        try:

            message_history = (
                get_message_history(
                    limit=50
                )
            )

        except Exception as exc:

            message_history = []

            st.error(
                "We couldn't load message history."
            )

            st.caption(
                f"DEV detail: {exc}"
            )

        if not message_history:

            st.info(
                "No message drafts have been saved yet."
            )

        else:

            history_rows = []

            for message in message_history:

                created_at = (
                    message.get(
                        "created_at"
                    )
                )

                created_label = (
                    created_at.strftime(
                        "%m/%d/%Y %I:%M %p"
                    )
                    if created_at is not None
                    else ""
                )

                saved_audiences = (
                    message.get(
                        "audiences"
                    )
                    or []
                )

                history_rows.append(
                    {
                        "ID":
                            message[
                                "message_id"
                            ],

                        "Created":
                            created_label,

                        "Audience":
                            ", ".join(
                                saved_audiences
                            ),

                        "Recipients":
                            message.get(
                                "recipient_count",
                                0,
                            ),

                        "Status":
                            str(
                                message.get(
                                    "status",
                                    ""
                                )
                            ).title(),

                        "Message":
                            message.get(
                                "message_text",
                                "",
                            ),
                    }
                )

            st.dataframe(
                pd.DataFrame(
                    history_rows
                ),
                hide_index=True,
                use_container_width=True,
                column_config={
                    "Message":
                        st.column_config.TextColumn(
                            "Message",
                            width="large",
                        ),
                },
            )

            message_by_id = {
                message[
                    "message_id"
                ]:
                    message

                for message
                in message_history
            }

            message_ids = list(
                message_by_id.keys()
            )

            selected_message_id = (
                st.selectbox(
                    "Inspect saved message",
                    message_ids,
                    format_func=(
                        lambda message_id: (
                            f"Draft #{message_id} · "
                            + ", ".join(
                                message_by_id[
                                    message_id
                                ].get(
                                    "audiences"
                                )
                                or []
                            )
                            + " · "
                            + str(
                                message_by_id[
                                    message_id
                                ].get(
                                    "recipient_count",
                                    0,
                                )
                            )
                            + " recipients"
                        )
                    ),
                    key="messaging_history_message_id",
                )
            )

            selected_message = (
                message_by_id[
                    selected_message_id
                ]
            )

            with st.container(
                border=True,
                key="messaging_history_detail",
            ):

                st.write(
                    "**Saved message**"
                )

                st.code(
                    selected_message.get(
                        "message_text",
                        "",
                    ),
                    language=None,
                )

                st.caption(
                    "Audience: "
                    + ", ".join(
                        selected_message.get(
                            "audiences"
                        )
                        or []
                    )
                    + "  •  "
                    + "Status: "
                    + str(
                        selected_message.get(
                            "status",
                            ""
                        )
                    ).title()
                )

                try:

                    saved_recipients = (
                        get_message_recipients(
                            selected_message_id
                        )
                    )

                except Exception as exc:

                    saved_recipients = []

                    st.error(
                        "We couldn't load the saved recipient snapshot."
                    )

                    st.caption(
                        f"DEV detail: {exc}"
                    )

                if saved_recipients:

                    saved_recipient_df = (
                        pd.DataFrame(
                            [
                                {
                                    "Household ID":
                                        recipient.get(
                                            "household_reference",
                                            "",
                                        ),

                                    "Contact":
                                        recipient.get(
                                            "contact_name",
                                            "",
                                        ),

                                    "Phone":
                                        recipient.get(
                                            "phone",
                                            "",
                                        ),

                                    "Using":
                                        recipient.get(
                                            "contact_source",
                                            "",
                                        ),

                                    "Children":
                                        recipient.get(
                                            "children",
                                            "",
                                        ),

                                    "Status":
                                        str(
                                            recipient.get(
                                                "status",
                                                "",
                                            )
                                        ).title(),
                                }

                                for recipient
                                in saved_recipients
                            ]
                        )
                    )

                    st.dataframe(
                        saved_recipient_df,
                        hide_index=True,
                        use_container_width=True,
                    )

                    saved_recipient_csv = (
                        saved_recipient_df
                        .to_csv(
                            index=False
                        )
                        .encode(
                            "utf-8-sig"
                        )
                    )

                    st.download_button(
                        "Download Saved Recipient Snapshot",
                        data=saved_recipient_csv,
                        file_name=(
                            "ascension-message-"
                            f"{selected_message_id}-"
                            "recipients.csv"
                        ),
                        mime="text/csv",
                        use_container_width=True,
                        key=(
                            "download_saved_message_"
                            f"{selected_message_id}"
                        ),
                    )

                if (
                    selected_message.get(
                        "status"
                    )
                    == "draft"
                ):

                    queue_col, delete_col = (
                        st.columns(2)
                    )

                    with queue_col:

                        if st.button(
                            "Queue Selected Draft",
                            type="primary",
                            use_container_width=True,
                            key=(
                                "queue_message_draft_"
                                f"{selected_message_id}"
                            ),
                        ):

                            try:

                                queued = (
                                    queue_message(
                                        selected_message_id
                                    )
                                )

                                if queued:

                                    st.success(
                                        f"Draft #{selected_message_id} "
                                        "is now queued for the DEV gateway. "
                                        "Nothing has been sent yet."
                                    )

                                    st.rerun()

                                else:

                                    st.warning(
                                        "That draft could not be queued."
                                    )

                            except Exception as exc:

                                st.error(
                                    "We couldn't queue this draft."
                                )

                                st.caption(
                                    f"DEV detail: {exc}"
                                )

                    with delete_col:

                        if st.button(
                            "Delete Selected Draft",
                            use_container_width=True,
                            key=(
                                "delete_message_draft_"
                                f"{selected_message_id}"
                            ),
                        ):

                            try:

                                deleted = (
                                    delete_message_draft(
                                        selected_message_id
                                    )
                                )

                                if deleted:

                                    st.success(
                                        f"Draft #{selected_message_id} deleted."
                                    )

                                    st.rerun()

                                else:

                                    st.warning(
                                        "That message could not be deleted. "
                                        "Only drafts may be removed."
                                    )

                            except Exception as exc:

                                st.error(
                                    "We couldn't delete this draft."
                                )

                                st.caption(
                                    f"DEV detail: {exc}"
                                )

                elif (
                    selected_message.get(
                        "status"
                    )
                    == "queued"
                ):

                    st.info(
                        "This message is locked in the DEV gateway queue. "
                        "The Pixel claims one household at a time. A claimed "
                        "household may be released back to the queue before "
                        "sending, and submitted messages are never "
                        "automatically retried."
                    )

    st.divider()

    st.header(
        "Sacramental Preparation"
    )

    first_communion_children = [
        child

        for child
        in roster

        if child.get(
            "receiving_first_communion_reconciliation",
            False,
        )
    ]

    confirmation_children = [
        child

        for child
        in roster

        if child.get(
            "receiving_confirmation",
            False,
        )
    ]

    follow_up_children = [
        child

        for child
        in roster

        if sacramental_follow_up_reasons(
            child
        )
    ]

    with st.expander(
        "First Reconciliation / "
        f"First Communion ({len(first_communion_children)})"
    ):

        if not first_communion_children:

            st.info(
                "No children are currently registered "
                "for First Reconciliation / First Communion."
            )

        else:

            rows = []

            for child in first_communion_children:

                rows.append(
                    {
                        "Child":
                            full_name(
                                child.get(
                                    "first_name"
                                ),
                                child.get(
                                    "middle_name"
                                ),
                                child.get(
                                    "last_name"
                                ),
                            ),

                        "Grade":
                            child[
                                "grade"
                            ],

                        "School":
                            child[
                                "school"
                            ],

                        "Baptized":
                            child.get(
                                "baptism_status"
                            )
                            or "",
                    }
                )

            st.dataframe(
                pd.DataFrame(
                    rows
                ),
                hide_index=True,
                use_container_width=True,
            )

            export_df = (
                build_export_dataframe(
                    first_communion_children
                )
            )

            st.download_button(
                "Download First Communion Roster",
                data=(
                    export_df
                    .to_csv(
                        index=False
                    )
                    .encode(
                        "utf-8-sig"
                    )
                ),
                file_name=(
                    "ascension-first-communion-"
                    "roster.csv"
                ),
                mime="text/csv",
                use_container_width=True,
            )

    with st.expander(
        f"Confirmation ({len(confirmation_children)})"
    ):

        if not confirmation_children:

            st.info(
                "No children are currently registered "
                "for Confirmation."
            )

        else:

            rows = []

            for child in confirmation_children:

                rows.append(
                    {
                        "Child":
                            full_name(
                                child.get(
                                    "first_name"
                                ),
                                child.get(
                                    "middle_name"
                                ),
                                child.get(
                                    "last_name"
                                ),
                            ),

                        "Grade":
                            child[
                                "grade"
                            ],

                        "School":
                            child[
                                "school"
                            ],

                        "Baptized":
                            child.get(
                                "baptism_status"
                            )
                            or "",

                        "Reconciliation":
                            child.get(
                                "first_reconciliation_status"
                            )
                            or "",

                        "Communion":
                            child.get(
                                "first_communion_status"
                            )
                            or "",
                    }
                )

            st.dataframe(
                pd.DataFrame(
                    rows
                ),
                hide_index=True,
                use_container_width=True,
            )

            export_df = (
                build_export_dataframe(
                    confirmation_children
                )
            )

            st.download_button(
                "Download Confirmation Roster",
                data=(
                    export_df
                    .to_csv(
                        index=False
                    )
                    .encode(
                        "utf-8-sig"
                    )
                ),
                file_name=(
                    "ascension-confirmation-"
                    "roster.csv"
                ),
                mime="text/csv",
                use_container_width=True,
            )

    with st.expander(
        f"Sacramental Follow-up Needed "
        f"({len(follow_up_children)})"
    ):

        if not follow_up_children:

            st.success(
                "No sacramental follow-up cases "
                "are currently flagged."
            )

        else:

            rows = []

            for child in follow_up_children:

                reasons = (
                    sacramental_follow_up_reasons(
                        child
                    )
                )

                rows.append(
                    {
                        "Child":
                            full_name(
                                child.get(
                                    "first_name"
                                ),
                                child.get(
                                    "middle_name"
                                ),
                                child.get(
                                    "last_name"
                                ),
                            ),

                        "Grade":
                            child[
                                "grade"
                            ],

                        "School":
                            child[
                                "school"
                            ],

                        "Follow-up Reason":
                            "; ".join(
                                reasons
                            ),

                        "Parent":
                            parent_name(
                                child.get(
                                    "parent_a_first_name"
                                ),
                                child.get(
                                    "parent_a_last_name"
                                ),
                            ),

                        "Email":
                            child.get(
                                "parent_a_email",
                                "",
                            ),

                        "Phone":
                            child.get(
                                "parent_a_phone",
                                "",
                            ),
                    }
                )

            st.dataframe(
                pd.DataFrame(
                    rows
                ),
                hide_index=True,
                use_container_width=True,
            )

            export_df = (
                build_export_dataframe(
                    follow_up_children
                )
            )

            st.download_button(
                "Download Follow-up List",
                data=(
                    export_df
                    .to_csv(
                        index=False
                    )
                    .encode(
                        "utf-8-sig"
                    )
                ),
                file_name=(
                    "ascension-sacramental-"
                    "follow-up.csv"
                ),
                mime="text/csv",
                use_container_width=True,
            )

    st.divider()

    st.header(
        "All Registrations"
    )

    with st.expander(
        "View / Export Full Registration Data"
    ):

        search_text = (
            st.text_input(
                "Search registrations",
                placeholder=(
                    "Child, parent, school, email, "
                    "or Household ID"
                ),
            )
        )

        search_value = (
            search_text
            .strip()
            .lower()
        )

        filtered = []

        for child in roster:

            child_name = full_name(
                child.get(
                    "first_name"
                ),
                child.get(
                    "middle_name"
                ),
                child.get(
                    "last_name"
                ),
            )

            parent_a = parent_name(
                child.get(
                    "parent_a_first_name"
                ),
                child.get(
                    "parent_a_last_name"
                ),
            )

            parent_b = parent_name(
                child.get(
                    "parent_b_first_name"
                ),
                child.get(
                    "parent_b_last_name"
                ),
            )

            searchable = " ".join(
                [
                    child_name,
                    parent_a,
                    parent_b,
                    child.get(
                        "school"
                    )
                    or "",
                    child.get(
                        "parent_a_email"
                    )
                    or "",
                    child.get(
                        "parent_b_email"
                    )
                    or "",
                    child.get(
                        "household_reference"
                    )
                    or "",
                    child.get(
                        "emergency_contact_name"
                    )
                    or "",
                    child.get(
                        "emergency_contact_relationship"
                    )
                    or "",
                    child.get(
                        "emergency_contact_phone"
                    )
                    or "",
                ]
            ).lower()

            if (
                search_value
                and search_value
                not in searchable
            ):

                continue

            filtered.append(
                child
            )

        st.caption(
            f"{len(filtered)} registration"
            f"{'' if len(filtered) == 1 else 's'} shown"
        )

        if filtered:

            simple_rows = []

            for child in filtered:

                simple_rows.append(
                    {
                        "Child":
                            full_name(
                                child.get(
                                    "first_name"
                                ),
                                child.get(
                                    "middle_name"
                                ),
                                child.get(
                                    "last_name"
                                ),
                            ),

                        "Grade":
                            child[
                                "grade"
                            ],

                        "School":
                            child[
                                "school"
                            ],

                        "Household ID":
                            child[
                                "household_reference"
                            ],

                        "Parent":
                            parent_name(
                                child.get(
                                    "parent_a_first_name"
                                ),
                                child.get(
                                    "parent_a_last_name"
                                ),
                            ),
                    }
                )

            st.dataframe(
                pd.DataFrame(
                    simple_rows
                ),
                hide_index=True,
                use_container_width=True,
            )

            full_export = (
                build_export_dataframe(
                    filtered
                )
            )

            st.download_button(
                "Download Current Registration Data",
                data=(
                    full_export
                    .to_csv(
                        index=False
                    )
                    .encode(
                        "utf-8-sig"
                    )
                ),
                file_name=(
                    "ascension-registration-data.csv"
                ),
                mime="text/csv",
                use_container_width=True,
            )

        else:

            st.info(
                "No registrations match your search."
            )

    if (
        st.session_state.admin_detail_child_id
        is not None
    ):

        selected_child = next(
            (
                child

                for child
                in roster

                if child[
                    "child_id"
                ]
                == st.session_state.admin_detail_child_id
            ),
            None,
        )

        if (
            selected_child
            is not None
        ):

            admin_child_detail_dialog(
                selected_child
            )

        else:

            clear_admin_child_detail()

    st.stop()


# ---------------------------------------------------------
# Registration complete
# ---------------------------------------------------------

if (
    st.session_state.submitted_household_id
    is not None
):

    editing_existing = (
        st.session_state.registration_mode
        == "existing"
    )

    household_reference = (
        st.session_state.submitted_household_reference
    )

    if LOGO_PATH.exists():

        logo_url = (
            image_to_data_url(
                LOGO_PATH
            )
        )

        st.html(
            dedent(
                f"""
                <div class="completion-logo">
                    <img
                        src="{logo_url}"
                        alt="Ascension Catholic Church"
                    >
                </div>
                """
            )
        )

    if editing_existing:

        st.html(
            dedent(
                """
                <div class="completion-kicker">
                    Registration Updated
                </div>

                <div class="completion-title">
                    Your changes have been saved.
                </div>

                <div class="completion-copy">
                    Your family's Faith Formation registration
                    has been updated successfully.
                </div>
                """
            )
        )

    else:

        st.html(
            dedent(
                """
                <div class="completion-kicker">
                    Registration Complete
                </div>

                <div class="completion-title">
                    You're all set!
                </div>

                <div class="completion-copy">
                    Thank you for registering your family for
                    Faith Formation at Ascension Catholic Church.
                </div>
                """
            )
        )


    # -----------------------------------------------------
    # Household ID / email confirmation
    # -----------------------------------------------------

    email_status_html = ""

    if (
        st.session_state.confirmation_email_sent
        is True
    ):

        email_address = escape_html(
            st.session_state.confirmation_email_address
            or ""
        )

        email_status_html = dedent(
            f"""
            <div class="completion-email-success">
                <span class="completion-email-check">✓</span>
                <span>
                    Confirmation sent to
                    <strong>{email_address}</strong>.
                    Your email also includes your Household ID.
                </span>
            </div>
            """
        )

    elif (
        st.session_state.confirmation_email_sent
        is False
    ):

        email_status_html = dedent(
            """
            <div class="completion-email-warning">
                Your registration was saved successfully,
                but we couldn't send the confirmation email.
                Please make a note of your Household ID before
                leaving this page.
            </div>
            """
        )

    if editing_existing:

        id_help = (
            "Keep this Household ID for future changes "
            "to your family's registration."
        )

    else:

        id_help = (
            "Keep this Household ID somewhere handy. "
            "You'll use it if you need to return later "
            "to update your registration."
        )

    summary_html = (
        dedent(
            f"""
            <div class="completion-summary">
                <div class="completion-id-label">
                    Your Household ID
                </div>

                <div class="completion-id-value">
                    {escape_html(household_reference)}
                </div>

                <div class="completion-id-help">
                    {escape_html(id_help)}
                </div>
            """
        )
        + email_status_html
        + "\n</div>"
    )

    st.html(
        summary_html
    )


    # -----------------------------------------------------
    # Next steps
    # -----------------------------------------------------

    if editing_existing:

        next_steps_html = dedent(
            """
            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                Your updated information is now on file.
            </div>

            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                Ascension staff can see your changes immediately.
            </div>

            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                You can return later using the same Household ID.
            </div>
            """
        )

    else:

        next_steps_html = dedent(
            """
            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                Your household and children are registered.
            </div>

            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                Ascension staff will review your registration.
            </div>

            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                We'll contact you if any sacramental follow-up
                is needed.
            </div>

            <div class="completion-next-item">
                <span class="completion-next-check">✓</span>
                You can return later using your Household ID
                if anything changes.
            </div>
            """
        )

    next_steps_card = (
        dedent(
            """
            <div class="completion-next">
                <div class="completion-next-title">
                    What happens next?
                </div>
            """
        )
        + next_steps_html
        + "\n</div>"
    )

    st.html(
        next_steps_card
    )


    # -----------------------------------------------------
    # Done
    # -----------------------------------------------------

    done_left, done_center, done_right = (
        st.columns(
            [
                1.3,
                1,
                1.3,
            ]
        )
    )

    with done_center:

        if st.button(
            "Done",
            type="primary",
            use_container_width=True,
        ):

            reset_public_registration_state()
            st.rerun()


    # -----------------------------------------------------
    # Footer
    # -----------------------------------------------------

    st.markdown(
        dedent(
            """
            <div class="completion-footer">
                Ascension Catholic Church<br>
                Faith Formation Registration
            </div>
            """
        ),
        unsafe_allow_html=True,
    )

    st.stop()


# ---------------------------------------------------------
# Entrance screen
# ---------------------------------------------------------

if (
    st.session_state.registration_mode
    is None
):

    if LOGO_PATH.exists():

        logo_url = (
            image_to_data_url(
                LOGO_PATH
            )
        )

        st.markdown(
            f"""
            <div class="landing-logo">
                <img
                    src="{logo_url}"
                    alt="Ascension Catholic Church"
                >
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown(
        """
        <div class="landing-parish">
            Ascension Catholic Church
        </div>

        <div class="landing-title">
            Faith Formation Registration
        </div>

        <div class="landing-welcome">
            Welcome! Use this portal to register your children
            for PSR, EDGE, Life Teen, and sacramental preparation
            at Ascension Catholic Church.
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.container(
        border=True,
        key="landing_new_card",
    ):

        st.markdown(
            """
            <div class="landing-section-label">
                New Registration
            </div>

            <div class="landing-action-title">
                Register your household
            </div>

            <div class="landing-action-description">
                Register your household and children for the
                upcoming faith formation year.
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(
            "Start a New Registration",
            type="primary",
            use_container_width=True,
        ):

            st.session_state.household = None
            st.session_state.children = []

            st.session_state.existing_household_id = None
            st.session_state.existing_household_reference = None

            st.session_state.confirmation_email_sent = None
            st.session_state.confirmation_email_address = None
            st.session_state.confirmation_email_error = None

            st.session_state.registration_mode = "new"

            clear_verification_state()
            clear_recovery_state()
            clear_admin_login_state()

            st.rerun()

    st.write("")

    with st.container(
        border=True,
        key="landing_returning_card",
    ):

        st.markdown(
            """
            <div class="landing-section-label">
                Returning Household
            </div>

            <div class="landing-action-title">
                Already registered?
            </div>

            <div class="landing-action-description">
                Return to your household to review information,
                add another child, or make changes.
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(
            "Return to Existing Household",
            use_container_width=True,
        ):

            clear_recovery_state()
            clear_admin_login_state()

            st.session_state.show_existing_dialog = True
            st.rerun()

    st.markdown(
        """
        <div class="landing-recovery">
            Can't find your Household ID?
        </div>
        """,
        unsafe_allow_html=True,
    )

    recovery_col_1, recovery_col_2, recovery_col_3 = (
        st.columns(
            [
                1.7,
                1,
                1.7,
            ]
        )
    )

    with recovery_col_2:

        if st.button(
            "Send My Household ID",
            type="tertiary",
            use_container_width=True,
        ):

            clear_verification_state()
            clear_admin_login_state()

            st.session_state.show_recovery_dialog = True
            st.rerun()

    st.divider()

    st.markdown(
        """
        <div class="landing-admin">
            Staff Administration
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.button(
        "Admin Login",
        type="tertiary",
        use_container_width=True,
    ):

        clear_verification_state()
        clear_recovery_state()

        st.session_state.show_admin_dialog = True
        st.rerun()

    if (
        st.session_state.show_existing_dialog
    ):

        existing_household_dialog()

    elif (
        st.session_state.show_recovery_dialog
    ):

        recover_household_id_dialog()

    elif (
        st.session_state.show_admin_dialog
    ):

        admin_login_dialog()

    st.stop()


# ---------------------------------------------------------
# Main registration workspace
# ---------------------------------------------------------

household = (
    st.session_state.household
)

children = (
    st.session_state.children
)

child_count = len(
    children
)

household_complete = (
    household_contact_complete(
        household
    )
)

registration_ready = (
    household_complete
    and child_count > 0
)


# ---------------------------------------------------------
# Return to start
# ---------------------------------------------------------

if st.button(
    "← Return to Start",
    type="tertiary",
):

    reset_public_registration_state()
    st.rerun()


# ---------------------------------------------------------
# Heading
# ---------------------------------------------------------

st.title(
    "Faith Formation Registration"
)

if (
    st.session_state.registration_mode
    == "existing"
):

    st.caption(
        f"Household ID: "
        f"{st.session_state.existing_household_reference}"
    )

    st.markdown(
        """
        <div class="registration-intro">
            Review your household information, update your
            children as needed, then review and save your changes.
        </div>
        """,
        unsafe_allow_html=True,
    )

else:

    st.markdown(
        """
        <div class="registration-intro">
            Complete your household information, add each child,
            then review everything before submitting.
        </div>
        """,
        unsafe_allow_html=True,
    )


# ---------------------------------------------------------
# Progress
# ---------------------------------------------------------

with st.container(
    border=True,
    key="progress_card",
):

    (
        progress_household,
        progress_children,
        progress_review,
    ) = st.columns(3)

    with progress_household:

        st.markdown(
            """
            <div class="progress-label">
                Step 1
            </div>

            <div class="progress-title">
                Household
            </div>
            """,
            unsafe_allow_html=True,
        )

        if household_complete:

            st.markdown(
                """
                <div class="progress-complete">
                    ✓ Complete
                </div>
                """,
                unsafe_allow_html=True,
            )

        elif household is not None:

            st.markdown(
                """
                <div class="progress-needed">
                    Action needed
                </div>
                """,
                unsafe_allow_html=True,
            )

        else:

            st.markdown(
                """
                <div class="progress-waiting">
                    Not started
                </div>
                """,
                unsafe_allow_html=True,
            )

    with progress_children:

        st.markdown(
            """
            <div class="progress-label">
                Step 2
            </div>

            <div class="progress-title">
                Children
            </div>
            """,
            unsafe_allow_html=True,
        )

        if child_count > 0:

            child_word = (
                "child"
                if child_count == 1
                else "children"
            )

            st.markdown(
                f"""
                <div class="progress-complete">
                    ✓ {child_count} {child_word} added
                </div>
                """,
                unsafe_allow_html=True,
            )

        else:

            st.markdown(
                """
                <div class="progress-waiting">
                    None added
                </div>
                """,
                unsafe_allow_html=True,
            )

    with progress_review:

        st.markdown(
            """
            <div class="progress-label">
                Step 3
            </div>

            <div class="progress-title">
                Review
            </div>
            """,
            unsafe_allow_html=True,
        )

        if registration_ready:

            st.markdown(
                """
                <div class="progress-complete">
                    ✓ Ready
                </div>
                """,
                unsafe_allow_html=True,
            )

        else:

            st.markdown(
                """
                <div class="progress-waiting">
                    Waiting
                </div>
                """,
                unsafe_allow_html=True,
            )


st.write("")


# ---------------------------------------------------------
# Household section
# ---------------------------------------------------------

st.markdown(
    """
    <div class="section-eyebrow">
        Step 1
    </div>
    """,
    unsafe_allow_html=True,
)

st.header(
    "Household"
)

if household is None:

    with st.container(
        border=True,
        key="household_empty_card",
    ):

        st.markdown(
            """
            <div class="empty-state-title">
                First, tell us about your household.
            </div>

            <div class="empty-state-copy">
                We'll use this information for registration
                communications and to identify your family.
                Household information must be completed before
                children can be added.
            </div>
            """,
            unsafe_allow_html=True,
        )

        if st.button(
            "Add Household Information",
            type="primary",
            use_container_width=True,
        ):

            household_dialog()

else:

    with st.container(
        border=True,
        key="household_card",
    ):

        if household_complete:

            st.caption(
                "✓ Household information complete"
            )

        else:

            st.warning(
                "Please add an emergency contact before "
                "reviewing or saving this registration."
            )

        parent_a_first = (
            household.get(
                "parent_a_first_name",
                "",
            ).strip()
        )

        parent_a_last = (
            household.get(
                "parent_a_last_name",
                "",
            ).strip()
        )

        parent_b_first = (
            household.get(
                "parent_b_first_name",
                "",
            ).strip()
        )

        parent_b_last = (
            household.get(
                "parent_b_last_name",
                "",
            ).strip()
        )

        parent_a = (
            f"{parent_a_first} "
            f"{parent_a_last}"
        ).strip()

        parent_b = (
            f"{parent_b_first} "
            f"{parent_b_last}"
        ).strip()

        if parent_b:

            if (
                parent_a_last
                and parent_b_last
                and parent_a_last.casefold()
                == parent_b_last.casefold()
            ):

                household_name = (
                    f"{parent_a_first} & "
                    f"{parent_b_first} "
                    f"{parent_a_last}"
                )

            else:

                household_name = (
                    f"{parent_a} & {parent_b}"
                )

        else:

            household_name = (
                parent_a
            )

        st.subheader(
            household_name
        )

        st.write(
            household[
                "address_line_1"
            ]
        )

        if household.get(
            "address_line_2"
        ):

            st.write(
                household[
                    "address_line_2"
                ]
            )

        st.write(
            f"{household['city']}, "
            f"{household['state']} "
            f"{household['zip_code']}"
        )

        st.write("")

        st.caption(
            f"{household['parent_a_email']} "
            f"• "
            f"{household['parent_a_phone']}"
        )

        if household.get(
            "emergency_contact_name"
        ):

            st.write("")

            st.caption(
                "Emergency contact"
            )

            st.write(
                f"{household['emergency_contact_name']} "
                f"• "
                f"{household['emergency_contact_relationship']} "
                f"• "
                f"{household['emergency_contact_phone']}"
            )

        if st.button(
            (
                "Edit Household"
                if household_complete
                else "Add Emergency Contact"
            ),
            key="edit_household",
        ):

            household_dialog()


st.divider()


# ---------------------------------------------------------
# Children section
# ---------------------------------------------------------

st.markdown(
    """
    <div class="section-eyebrow">
        Step 2
    </div>
    """,
    unsafe_allow_html=True,
)

if child_count == 0:

    st.header(
        "Children"
    )

else:

    st.header(
        f"Children · {child_count} added"
    )


if household is None:

    st.caption(
        "Complete your household information first. "
        "You'll then be able to add children."
    )

else:

    if not children:

        with st.container(
            border=True,
            key="children_empty_card",
        ):

            st.markdown(
                """
                <div class="empty-state-title">
                    Next, add each child you're registering.
                </div>

                <div class="empty-state-copy">
                    You'll enter basic school information and
                    indicate whether the child is preparing for
                    First Reconciliation / First Communion or
                    Confirmation this year.
                </div>
                """,
                unsafe_allow_html=True,
            )

            if st.button(
                "＋ Add a Child",
                type="primary",
                use_container_width=True,
                key="first_add_child",
            ):

                child_dialog()

    else:

        for index, child in enumerate(
            children
        ):

            with st.container(
                border=True,
                key=f"child_card_{index}",
            ):

                child_name = full_name(
                    child.get(
                        "first_name"
                    ),
                    child.get(
                        "middle_name"
                    ),
                    child.get(
                        "last_name"
                    ),
                )

                age = calculate_age(
                    child[
                        "date_of_birth"
                    ]
                )

                st.subheader(
                    child_name
                )

                st.caption(
                    f"{child_grade_label(child['grade'])} "
                    f"• {child['school']} "
                    f"• Age {age}"
                )

                preparation = (
                    sacrament_preparation_labels(
                        child
                    )
                )

                if preparation:

                    st.caption(
                        "Sacrament preparation: "
                        + ", ".join(
                            preparation
                        )
                    )

                follow_up_reasons = (
                    sacramental_follow_up_reasons(
                        child
                    )
                )

                if follow_up_reasons:

                    st.warning(
                        "Sacramental follow-up needed"
                    )

                edit_col, remove_col = (
                    st.columns(2)
                )

                with edit_col:

                    if st.button(
                        "Edit",
                        key=(
                            f"edit_child_"
                            f"{index}"
                        ),
                        use_container_width=True,
                    ):

                        child_dialog(
                            index
                        )

                with remove_col:

                    if st.button(
                        "Remove",
                        key=(
                            f"remove_child_"
                            f"{index}"
                        ),
                        use_container_width=True,
                    ):

                        st.session_state.children.pop(
                            index
                        )

                        st.rerun()

        if st.button(
            "＋ Add Another Child",
            type="primary",
            use_container_width=True,
        ):

            child_dialog()


st.divider()


# ---------------------------------------------------------
# Review section
# ---------------------------------------------------------

st.markdown(
    """
    <div class="section-eyebrow">
        Step 3
    </div>
    """,
    unsafe_allow_html=True,
)

st.header(
    "Review & Submit"
)

if registration_ready:

    with st.container(
        border=True,
        key="review_ready_card",
    ):

        st.markdown(
            """
            <div class="review-ready-title">
                Everything looks ready.
            </div>
            """,
            unsafe_allow_html=True,
        )

        child_word = (
            "child"
            if child_count == 1
            else "children"
        )

        st.markdown(
            f"""
            <div class="review-check">
                ✓ Household information complete<br>
                ✓ Emergency contact provided<br>
                ✓ {child_count} {child_word} added
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.write("")

        if (
            st.session_state.registration_mode
            == "existing"
        ):

            review_button_text = (
                "Review Changes"
            )

        else:

            review_button_text = (
                "Review Registration"
            )

        if st.button(
            review_button_text,
            type="primary",
            use_container_width=True,
        ):

            review_dialog()

else:

    with st.container(
        border=True,
        key="review_pending_card",
    ):

        st.markdown(
            """
            <div class="review-ready-title">
                A little more to go.
            </div>
            """,
            unsafe_allow_html=True,
        )

        if household is None:

            st.write(
                "○ Complete household information"
            )

        elif not household_complete:

            st.write(
                "○ Add an emergency contact"
            )

        else:

            st.write(
                "✓ Household information complete"
            )

        if child_count == 0:

            st.write(
                "○ Add at least one child"
            )

        else:

            st.write(
                f"✓ {child_count} "
                f"{'child' if child_count == 1 else 'children'} added"
            )

        st.write("")

        st.button(
            "Review Registration",
            disabled=True,
            use_container_width=True,
        )


# ---------------------------------------------------------
# Privacy note
# ---------------------------------------------------------

st.markdown(
    """
    <div class="privacy-note">
        Registration information is collected by
        Ascension Catholic Church for faith formation,
        sacramental preparation, and ministry administration.
    </div>
    """,
    unsafe_allow_html=True,
)